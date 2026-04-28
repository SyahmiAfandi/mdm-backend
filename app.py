from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from flask import Response
from werkzeug.utils import secure_filename
from flask import send_file
from datetime import datetime, timedelta
from supabase import create_client, Client
import xlsxwriter
import io
import uuid, pickle, os
from copy import copy
import pandas as pd
import numpy as np
import glob
import json
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import os, sys
from pathlib import Path
import argparse
from datetime import datetime
import time
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font
from openpyxl.formatting.rule import FormulaRule
from io import BytesIO
import firebase_admin
from firebase_admin import credentials, firestore
from dotenv import load_dotenv
load_dotenv()

app = Flask(__name__)
CORS(app)

_firestore_client = None

def init_firestore():
    """
    Initialize Firebase Admin Firestore client.
    Recommended env var:
      - FIREBASE_SERVICE_ACCOUNT_JSON  (full JSON content as string)
        OR
      - FIREBASE_SERVICE_ACCOUNT_PATH  (path to json file)
    """
    global _firestore_client
    if _firestore_client is not None:
        return _firestore_client

    if not firebase_admin._apps:
        sa_json = os.getenv("FIREBASE_SERVICE_ACCOUNT_JSON", "").strip()
        sa_path = os.getenv("FIREBASE_SERVICE_ACCOUNT_PATH", "").strip()

        if sa_json:
            cred = credentials.Certificate(json.loads(sa_json))
        elif sa_path:
            cred = credentials.Certificate(sa_path)
        else:
            raise RuntimeError(
                "Missing Firebase credentials. Set FIREBASE_SERVICE_ACCOUNT_JSON or FIREBASE_SERVICE_ACCOUNT_PATH."
            )

        firebase_admin.initialize_app(cred)

    _firestore_client = firestore.client()
    return _firestore_client

_supabase_client = None

def init_supabase() -> Client:
    global _supabase_client
    if _supabase_client is not None:
        return _supabase_client

    url = os.getenv("SUPABASE_URL")
    key = (
        os.getenv("SUPABASE_SECRET_KEY")
        or os.getenv("SUPABASE_SERVICE_ROLE_KEY")
        or os.getenv("SUPABASE_KEY")
    )
    if not url or not key:
        raise RuntimeError(
            "Missing Supabase credentials. Set SUPABASE_URL and "
            "SUPABASE_SECRET_KEY (or SUPABASE_SERVICE_ROLE_KEY / SUPABASE_KEY)."
        )

    if key.startswith("sb_publishable_"):
        raise RuntimeError(
            "Invalid backend Supabase key. Use a secret key (sb_secret_...), not the publishable browser key."
        )

    _supabase_client = create_client(url, key)
    return _supabase_client

supabase = init_supabase()

MONTH_NAMES = [
    "January","February","March","April","May","June",
    "July","August","September","October","November","December"
]

def month_to_mm(month_value: str) -> str:
    v = str(month_value or "").strip()

    # Accept numeric months too: "2", "02"
    if v.isdigit():
        n = int(v)
        if 1 <= n <= 12:
            return f"{n:02d}"
        raise ValueError(f"Invalid numeric month '{v}'. Expected 1..12")

    # Accept month names: "February"
    if v not in MONTH_NAMES:
        raise ValueError(f"Invalid month '{v}'. Expected one of {MONTH_NAMES} or numeric 1..12")
    idx = MONTH_NAMES.index(v) + 1
    return f"{idx:02d}"

def month_to_number(month_value: str) -> int:
    v = str(month_value or "").strip()

    # If numeric already ("2", "02")
    if v.isdigit():
        n = int(v)
        if 1 <= n <= 12:
            return n
        raise ValueError(f"Invalid numeric month '{v}'. Expected 1..12")

    # If month name ("February")
    if v not in MONTH_NAMES:
        raise ValueError(f"Invalid month '{v}'. Expected one of {MONTH_NAMES} or numeric 1..12")

    return MONTH_NAMES.index(v) + 1

def make_period_id(year: str, month_name: str) -> str:
    # Example: 2026 + February -> "2026-02"
    return f"{str(year).strip()}-{month_to_mm(month_name)}"

def norm_status_ui(v: str) -> str:
    # UI sends "Match" / "Mismatch"
    s = (v or "").strip().lower()
    if s == "match":
        return "match"
    if s == "mismatch":
        return "mismatch"
    if s in ["no data", "nodata", "no_data"]:
        return "no_data"
    return s  # fallback

def make_cell_id(period_id: str, business_type: str, distributor_code: str, report_type_id: str) -> str:
    return f"{period_id}__{business_type}__{distributor_code}__{report_type_id}"

def first_supabase_row(response):
    data = getattr(response, "data", None)
    if isinstance(data, list):
        return data[0] if data else None
    return data if isinstance(data, dict) else None

def mirror_legacy_recon_cell(cell_id: str, payload: dict):
    """
    Best-effort mirror for older pages that still read the legacy camelCase table.
    This should never block the primary recon_cells write.
    """
    try:
        prev = first_supabase_row(
            supabase.table("reconCells").select("id").eq("id", cell_id).limit(1).execute()
        )
        if prev:
            supabase.table("reconCells").update(
                {k: v for k, v in payload.items() if k != "id"}
            ).eq("id", cell_id).execute()
        else:
            supabase.table("reconCells").insert(payload).execute()
    except Exception as exc:
        app.logger.warning("Skipping legacy reconCells mirror for %s: %s", cell_id, exc)

def get_legacy_recon_cell(cell_id: str):
    try:
        return first_supabase_row(
            supabase.table("reconCells")
            .select("id,status,reconsNo,createdAt,createdBy,updatedBy")
            .eq("id", cell_id)
            .limit(1)
            .execute()
        )
    except Exception as exc:
        app.logger.warning("Could not read legacy reconCells row for %s: %s", cell_id, exc)
        return None

DSS_REPORT_TABLE = "daily_sales_summary_reports"
DSS_RECON_TEMPLATE_PATH = Path(__file__).resolve().parent / "HPCIC DSS Recon Summary.xlsx"
DSS_TEMPLATE_CONFIG_PATH = Path(__file__).resolve().parent / "dss_template_config.json"
DSS_METRIC_COLUMNS = [
    ("sales_qty_cs", "Sale Qty CS"),
    ("sale_qty_pc", "Sale Qty PC"),
    ("free_total_qty", "Free Total Qty"),
    ("gsv_amount", "GSV(Amount)"),
    ("niv_total", "NIV(Net Invoice Value)"),
    ("sales_turn_over", "Sales Turnover"),
]

def normalize_distributor_code(value):
    return (
        str(value or "")
        .replace(".0", "")
        .strip()
    )

def safe_metric_number(value):
    numeric = pd.to_numeric(value, errors="coerce")
    if pd.isna(numeric):
        return 0.0
    return round(float(numeric), 4)

def pick_first_present(*values):
    for value in values:
        if value is None:
            continue
        if isinstance(value, str) and value.strip() == "":
            continue
        return value
    return ""

def normalize_dss_export_row(row):
    row = row or {}
    return {
        "distributor_code": normalize_distributor_code(
            pick_first_present(
                row.get("distributor_code"),
                row.get("distributorCode"),
                row.get("Distributor"),
                row.get("Distributor Code"),
                row.get("distributor"),
            )
        ),
        "distributor_name": str(
            pick_first_present(
                row.get("distributor_name"),
                row.get("distributorName"),
                row.get("Distributor Name"),
                row.get("Name"),
                row.get("name"),
            ) or ""
        ).strip(),
        "csdp_sales_qty_cs": safe_metric_number(
            pick_first_present(
                row.get("csdp_sales_qty_cs"),
                row.get("csdpSalesQtyCs"),
                row.get("CSDP Sales Qty CS"),
                row.get("CSDP_Sales Qty CS"),
            )
        ),
        "osdp_sales_qty_cs": safe_metric_number(
            pick_first_present(
                row.get("osdp_sales_qty_cs"),
                row.get("osdpSalesQtyCs"),
                row.get("OSDP Sales Qty CS"),
                row.get("OSDP_Sales Qty CS"),
            )
        ),
        "csdp_sale_qty_pc": safe_metric_number(
            pick_first_present(
                row.get("csdp_sale_qty_pc"),
                row.get("csdpSaleQtyPc"),
                row.get("CSDP Sale Qty PC"),
                row.get("CSDP_Sale Qty PC"),
            )
        ),
        "osdp_sale_qty_pc": safe_metric_number(
            pick_first_present(
                row.get("osdp_sale_qty_pc"),
                row.get("osdpSaleQtyPc"),
                row.get("OSDP Sale Qty PC"),
                row.get("OSDP_Sale Qty PC"),
            )
        ),
        "csdp_free_total_qty": safe_metric_number(
            pick_first_present(
                row.get("csdp_free_total_qty"),
                row.get("csdpFreeTotalQty"),
                row.get("CSDP Free Total Qty"),
                row.get("CSDP_Free Total Qty"),
            )
        ),
        "osdp_free_total_qty": safe_metric_number(
            pick_first_present(
                row.get("osdp_free_total_qty"),
                row.get("osdpFreeTotalQty"),
                row.get("OSDP Free Total Qty"),
                row.get("OSDP_Free Total Qty"),
            )
        ),
        "csdp_gsv_amount": safe_metric_number(
            pick_first_present(
                row.get("csdp_gsv_amount"),
                row.get("csdpGsvAmount"),
                row.get("CSDP GSV"),
                row.get("CSDP GSV Amount"),
                row.get("CSDP_GSV"),
                row.get("CSDP_GSV Amount"),
            )
        ),
        "osdp_gsv_amount": safe_metric_number(
            pick_first_present(
                row.get("osdp_gsv_amount"),
                row.get("osdpGsvAmount"),
                row.get("OSDP GSV"),
                row.get("OSDP GSV Amount"),
                row.get("OSDP_GSV"),
                row.get("OSDP_GSV Amount"),
            )
        ),
        "csdp_niv_total": safe_metric_number(
            pick_first_present(
                row.get("csdp_niv_total"),
                row.get("csdpNivTotal"),
                row.get("CSDP NIV"),
                row.get("CSDP NIV Total"),
                row.get("CSDP_NIV"),
                row.get("CSDP_NIV Total"),
            )
        ),
        "osdp_niv_total": safe_metric_number(
            pick_first_present(
                row.get("osdp_niv_total"),
                row.get("osdpNivTotal"),
                row.get("OSDP NIV"),
                row.get("OSDP NIV Total"),
                row.get("OSDP_NIV"),
                row.get("OSDP_NIV Total"),
            )
        ),
        "csdp_sales_turn_over": safe_metric_number(
            pick_first_present(
                row.get("csdp_sales_turn_over"),
                row.get("csdpSalesTurnOver"),
                row.get("CSDP Sales Turn Over"),
                row.get("CSDP_Sales Turn Over"),
            )
        ),
        "osdp_sales_turn_over": safe_metric_number(
            pick_first_present(
                row.get("osdp_sales_turn_over"),
                row.get("osdpSalesTurnOver"),
                row.get("OSDP Sales Turn Over"),
                row.get("OSDP_Sales Turn Over"),
            )
        ),
    }

def distributor_is_active(row):
    active_value = row.get("active")
    if isinstance(active_value, bool):
        is_active = active_value
    elif active_value is None:
        is_active = True
    else:
        active_text = str(active_value).strip().lower()
        is_active = active_text not in {"false", "0", "inactive", "no", "n"}

    status_text = str(
        pick_first_present(
            row.get("status"),
            row.get("activeLabel"),
        ) or ""
    ).strip().lower()
    if status_text in {"inactive", "disabled"}:
        return False
    return is_active

def normalize_country_code(value):
    return str(value or "").strip().upper()

def extract_distributor_countries(row):
    raw_values = []
    countries = row.get("countries")
    if isinstance(countries, list):
        raw_values.extend(countries)

    raw_values.extend([
        row.get("country"),
        row.get("countryCode"),
        row.get("country_code"),
    ])

    seen = set()
    normalized = []
    for value in raw_values:
        code = normalize_country_code(value)
        if not code or code in seen:
            continue
        seen.add(code)
        normalized.append(code)
    return normalized

def get_active_master_distributor_rows(business_type="HPC"):
    wanted_business = str(business_type or "").strip().upper() or "HPC"
    response = supabase.table("master_distributors").select("*").execute()
    rows = getattr(response, "data", None) or []

    active_rows = []
    for raw_row in rows:
        row_business = str(
            pick_first_present(
                raw_row.get("business_type"),
                raw_row.get("businessType"),
            ) or ""
        ).strip().upper()
        if row_business != wanted_business:
            continue
        if not distributor_is_active(raw_row):
            continue

        active_rows.append(raw_row)

    return active_rows

def get_default_dss_template_allowed_countries(available_codes):
    unique_codes = []
    seen = set()
    for value in available_codes or []:
        code = normalize_country_code(value)
        if not code or code in seen:
            continue
        seen.add(code)
        unique_codes.append(code)

    default_codes = [
        code for code in unique_codes
        if code not in {"SG", "USG"} and not code.endswith("SG")
    ]
    return default_codes or unique_codes

def load_dss_template_country_config():
    if not DSS_TEMPLATE_CONFIG_PATH.exists():
        return {}

    try:
        payload = json.loads(DSS_TEMPLATE_CONFIG_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}

    return payload if isinstance(payload, dict) else {}

def save_dss_template_country_config(config):
    payload = config if isinstance(config, dict) else {}
    DSS_TEMPLATE_CONFIG_PATH.write_text(
        json.dumps(payload, indent=2, sort_keys=True),
        encoding="utf-8",
    )

def get_dss_template_country_options(business_type="HPC"):
    counts = {}
    for raw_row in get_active_master_distributor_rows(business_type=business_type):
        for code in extract_distributor_countries(raw_row):
            counts[code] = counts.get(code, 0) + 1

    options = [
        {"code": code, "distributorCount": counts[code]}
        for code in sorted(counts.keys())
    ]
    return options

def resolve_dss_template_country_config(business_type="HPC"):
    business_key = str(business_type or "").strip().upper() or "HPC"
    available_options = get_dss_template_country_options(business_type=business_key)
    available_codes = [option["code"] for option in available_options]

    raw_config = load_dss_template_country_config()
    business_config = raw_config.get(business_key)
    if not isinstance(business_config, dict):
        business_config = {}

    configured_countries = business_config.get("allowedCountries")
    if isinstance(configured_countries, list):
        allowed_countries = [
            normalize_country_code(value)
            for value in configured_countries
            if normalize_country_code(value) in available_codes
        ]
        source = "saved"
    else:
        allowed_countries = get_default_dss_template_allowed_countries(available_codes)
        source = "default"

    return {
        "businessType": business_key,
        "allowedCountries": allowed_countries,
        "availableCountries": available_options,
        "source": source,
        "updatedAt": business_config.get("updatedAt"),
        "updatedBy": business_config.get("updatedBy"),
    }

def update_dss_template_country_config(business_type="HPC", allowed_countries=None, updated_by=""):
    business_key = str(business_type or "").strip().upper() or "HPC"
    current = load_dss_template_country_config()
    if not isinstance(current, dict):
        current = {}

    resolved = resolve_dss_template_country_config(business_type=business_key)
    available_codes = {option["code"] for option in resolved["availableCountries"]}
    cleaned = []
    seen = set()
    for value in allowed_countries or []:
        code = normalize_country_code(value)
        if not code or code in seen or code not in available_codes:
            continue
        seen.add(code)
        cleaned.append(code)

    current[business_key] = {
        "allowedCountries": cleaned,
        "updatedAt": datetime.utcnow().isoformat() + "Z",
        "updatedBy": str(updated_by or "").strip() or None,
    }
    save_dss_template_country_config(current)
    return resolve_dss_template_country_config(business_type=business_key)

def get_active_master_distributors(business_type="HPC", allowed_countries=None):
    allowed_country_set = None
    if allowed_countries is not None:
        allowed_country_set = {
            normalize_country_code(value)
            for value in allowed_countries
            if normalize_country_code(value)
        }

    active_by_code = {}
    for raw_row in get_active_master_distributor_rows(business_type=business_type):
        distributor_countries = extract_distributor_countries(raw_row)
        if allowed_country_set is not None:
            if not distributor_countries:
                continue
            if not any(code in allowed_country_set for code in distributor_countries):
                continue

        code = normalize_distributor_code(
            pick_first_present(
                raw_row.get("code"),
                raw_row.get("distributor_code"),
            )
        )
        name = str(
            pick_first_present(
                raw_row.get("name"),
                raw_row.get("distributor_name"),
            ) or ""
        ).strip()
        if not code:
            continue

        active_by_code[code] = {
            "code": code,
            "name": name,
            "countries": distributor_countries,
        }

    active_rows = list(active_by_code.values())
    active_rows.sort(key=lambda row: (row["code"], row["name"]))
    return active_rows

def extract_missing_schema_column(message, table_name=""):
    raw_message = str(message or "")
    marker = "Could not find the '"
    start = raw_message.find(marker)
    if start < 0:
        return ""

    start += len(marker)
    end = raw_message.find("' column", start)
    if end < 0:
        return ""

    column_name = raw_message[start:end].strip()
    if not column_name:
        return ""

    if table_name and f"of '{table_name}'" not in raw_message:
        return ""

    return column_name

def upsert_payloads_with_schema_retry(
    table_name: str,
    payloads: list,
    on_conflict: str = "id",
    chunk_size: int = 500,
):
    trimmed_payloads = [dict(payload) for payload in payloads]
    dropped_columns = []

    while True:
        try:
            for start in range(0, len(trimmed_payloads), chunk_size):
                supabase.table(table_name).upsert(
                    trimmed_payloads[start:start + chunk_size],
                    on_conflict=on_conflict,
                ).execute()
            return dropped_columns
        except Exception as exc:
            missing_column = extract_missing_schema_column(str(exc), table_name=table_name)
            if (
                not missing_column
                or missing_column == on_conflict
                or missing_column in dropped_columns
                or not any(missing_column in payload for payload in trimmed_payloads)
            ):
                raise

            dropped_columns.append(missing_column)
            trimmed_payloads = [
                {key: value for key, value in payload.items() if key != missing_column}
                for payload in trimmed_payloads
            ]
            app.logger.warning(
                "Dropping unsupported column '%s' while upserting '%s'.",
                missing_column,
                table_name,
            )

def build_daily_sales_export_frame(raw_df: pd.DataFrame) -> pd.DataFrame:
    if raw_df is None or raw_df.empty or "Distributor" not in raw_df.columns:
        return pd.DataFrame(columns=["Distributor", "Distributor Name", *[label for _, label in DSS_METRIC_COLUMNS]])

    df = raw_df.copy()
    df.columns = [str(col).strip() for col in df.columns]

    if "Sales Turn Over" in df.columns and "Sales Turnover" not in df.columns:
        df.rename(columns={"Sales Turn Over": "Sales Turnover"}, inplace=True)

    df["Distributor"] = (
        df["Distributor"]
        .astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.strip()
        .replace(["nan", "NaN", "None"], "")
    )

    if "Distributor Name" not in df.columns:
        df["Distributor Name"] = ""

    df["Distributor Name"] = (
        df["Distributor Name"]
        .fillna("")
        .astype(str)
        .str.strip()
        .replace(["nan", "NaN", "None"], "")
    )

    df = df[df["Distributor"] != ""].copy()
    if df.empty:
        return pd.DataFrame(columns=["Distributor", "Distributor Name", *[label for _, label in DSS_METRIC_COLUMNS]])

    metric_labels = []
    for _, label in DSS_METRIC_COLUMNS:
        if label in df.columns:
            df[label] = pd.to_numeric(df[label], errors="coerce").fillna(0.0)
            metric_labels.append(label)

    if not metric_labels:
        return pd.DataFrame(columns=["Distributor", "Distributor Name", *[label for _, label in DSS_METRIC_COLUMNS]])

    preferred_names = (
        df.assign(_name_rank=df["Distributor Name"].eq("").astype(int))
        .sort_values(["Distributor", "_name_rank"], kind="stable")
        [["Distributor", "Distributor Name"]]
        .drop_duplicates(subset=["Distributor"], keep="first")
    )

    grouped = df.groupby("Distributor", as_index=False)[metric_labels].sum()
    grouped = grouped.merge(preferred_names, on="Distributor", how="left")

    for _, label in DSS_METRIC_COLUMNS:
        if label not in grouped.columns:
            grouped[label] = 0.0

    return grouped[["Distributor", "Distributor Name", *[label for _, label in DSS_METRIC_COLUMNS]]]

def build_daily_sales_report_rows(osdp_df: pd.DataFrame, pbi_df: pd.DataFrame):
    osdp_export = build_daily_sales_export_frame(osdp_df)
    pbi_export = build_daily_sales_export_frame(pbi_df)

    merged = pd.merge(
        osdp_export,
        pbi_export,
        on="Distributor",
        how="outer",
        suffixes=("_osdp", "_csdp"),
    )

    if merged.empty:
        return []

    distributor_name = (
        merged.get("Distributor Name_csdp", pd.Series("", index=merged.index))
        .fillna("")
        .astype(str)
        .str.strip()
    )
    distributor_name = distributor_name.mask(
        distributor_name == "",
        merged.get("Distributor Name_osdp", pd.Series("", index=merged.index)).fillna("").astype(str).str.strip(),
    )

    merged["Distributor Name"] = distributor_name
    merged["Distributor"] = (
        merged["Distributor"]
        .astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.strip()
    )
    merged = merged.sort_values(by="Distributor", kind="stable").reset_index(drop=True)

    rows = []
    for _, row in merged.iterrows():
        distributor_code = normalize_distributor_code(row.get("Distributor"))
        if not distributor_code:
            continue

        payload = {
            "distributor_code": distributor_code,
            "distributor_name": str(row.get("Distributor Name", "") or "").strip(),
        }

        for metric_key, metric_label in DSS_METRIC_COLUMNS:
            osdp_value = safe_metric_number(row.get(f"{metric_label}_osdp", 0))
            csdp_value = safe_metric_number(row.get(f"{metric_label}_csdp", 0))
            payload[f"osdp_{metric_key}"] = osdp_value
            payload[f"csdp_{metric_key}"] = csdp_value

        rows.append(payload)

    return rows

##----------XLS XLSX FILE------------
# -------- Excel reader that supports mixed .xls + .xlsx (even mislabeled) --------
XLS_OLE_SIG = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"  # true legacy .xls (OLE2)
ZIP_SIG = b"PK"  # .xlsx is a ZIP container

def read_excel_mixed(file_storage, **kwargs):
    """
    Read Flask FileStorage that may be .xls or .xlsx.
    Works even if the filename extension is wrong.
    """
    data = file_storage.read()   # bytes
    bio = BytesIO(data)
    head = data[:8]

    if head.startswith(ZIP_SIG):
        return pd.read_excel(bio, engine="openpyxl", **kwargs)

    if head.startswith(XLS_OLE_SIG):
        return pd.read_excel(bio, engine="xlrd", **kwargs)

    raise ValueError("Unsupported/unknown Excel format (not xls/xlsx).")

####-------------Promo Preview-------------

# Allow your Vite dev server to talk to Flask



UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

RESULTS_DIR = 'reconcile_results'
os.makedirs(RESULTS_DIR, exist_ok=True)

# Future behavior: avoid silent downcasting warnings everywhere (optional)
pd.set_option('future.no_silent_downcasting', True)


def find_credentials(cli_arg: str | None = None) -> str:
    # 1) CLI argument wins
    if cli_arg:
        p = Path(cli_arg)
        if p.exists():
            return str(p)

    # 2) Environment variable
    env_p = os.getenv("GOOGLE_APPLICATION_CREDENTIALS")
    if env_p and Path(env_p).exists():
        return env_p

    # 3) Next to the EXE (portable case)
    if getattr(sys, "frozen", False):
        exe_dir = Path(sys.executable).parent
        cand = exe_dir / "credentials.json"
        if cand.exists():
            return str(cand)

    # 4) PyInstaller _MEIPASS (only if you decided to bundle it)
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        cand = Path(sys._MEIPASS) / "credentials.json"
        if cand.exists():
            return str(cand)

    # 5) Dev: next to app.py
    cand = Path(__file__).parent / "credentials.json"
    if cand.exists():
        return str(cand)

    raise FileNotFoundError(
        "credentials.json not found. Pass --creds PATH, set "
        "GOOGLE_APPLICATION_CREDENTIALS, or place credentials.json next to the EXE/app.py."
    )
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--port", type=int, default=int(os.getenv("PORT", 5000)))
    parser.add_argument("--creds", type=str, default=None)  # <— NEW
    args = parser.parse_args()

    # ---- BEFORE you create the Google client, resolve the credentials file
    cred_path = find_credentials(args.creds)

    # Example using oauth2client (what your traceback shows):
    from oauth2client.service_account import ServiceAccountCredentials
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets']  # adjust
    cred_path = "credentials.json"  # the path to the JSON file you downloaded
    CREDS = ServiceAccountCredentials.from_json_keyfile_name(cred_path, scopes=SCOPES)

    # ... use CREDS in your Google API code ...

    app.run(host="127.0.0.1", port=args.port, debug=False, use_reloader=False)



def convert_np(obj):
    if isinstance(obj, (np.integer, np.int64)):
        return int(obj)
    if isinstance(obj, (np.floating, np.float64)):
        return float(obj)
    return str(obj)

@app.route("/health", methods=["GET"])
def health():
    t0 = time.time()
    # TODO: put quick checks here if you want (e.g., DB ping, file system, etc.)
    latency_ms = int((time.time() - t0) * 1000)
    return jsonify({
        "service": "API Service",
        "status": "UP",                   # or "DEGRADED"/"DOWN" if checks fail
        "latencyMs": latency_ms,
        "updatedAt": datetime.utcnow().isoformat() + "Z",
    })

def auth_response_to_user(auth_response):
    if auth_response is None:
        return None

    user = getattr(auth_response, "user", None)
    if user is not None:
        return user

    if hasattr(auth_response, "model_dump"):
        payload = auth_response.model_dump()
        if isinstance(payload, dict):
            return payload.get("user") or ((payload.get("data") or {}).get("user"))

    if isinstance(auth_response, dict):
        return auth_response.get("user") or ((auth_response.get("data") or {}).get("user"))

    return None

def get_auth_user_field(user, field_name):
    if user is None:
        return None
    if isinstance(user, dict):
        return user.get(field_name)
    return getattr(user, field_name, None)

def require_admin_request_user():
    auth_header = str(request.headers.get("Authorization") or "").strip()
    if not auth_header.lower().startswith("bearer "):
        return None, (jsonify({"error": "Missing Authorization bearer token"}), 401)

    token = auth_header[7:].strip()
    if not token:
        return None, (jsonify({"error": "Missing Authorization bearer token"}), 401)

    try:
        auth_response = supabase.auth.get_user(token)
        auth_user = auth_response_to_user(auth_response)
        uid = get_auth_user_field(auth_user, "id")

        if not uid:
            return None, (jsonify({"error": "Invalid or expired session"}), 401)

        role_row = first_supabase_row(
            supabase.table("user_roles").select("role").eq("id", uid).limit(1).execute()
        )
        role = (role_row or {}).get("role")

        if role != "admin":
            return None, (jsonify({"error": "Admin access required"}), 403)

        return {
            "id": uid,
            "email": get_auth_user_field(auth_user, "email"),
            "role": role,
        }, None
    except Exception as exc:
        app.logger.warning("Failed to authorize admin request: %s", exc)
        return None, (jsonify({"error": "Invalid or expired session"}), 401)

def license_date_to_valid_until(ymd):
    if not ymd:
        return None

    base = datetime.strptime(str(ymd).strip(), "%Y-%m-%d")
    next_day = base + timedelta(days=1)
    return next_day.isoformat() + "Z"

@app.route("/admin/reset-password", methods=["POST"])
def admin_reset_password():
    """
    Admin-only endpoint to reset a user's password using Supabase Admin Auth.
    Expects JSON: { "user_id": "...", "new_password": "..." }
    """
    try:
        data = request.get_json()
        user_id = data.get("user_id")
        new_password = data.get("new_password")

        if not user_id or not new_password:
            return jsonify({"error": "user_id and new_password are required"}), 400

        # Update user password using Service Role (Admin) privileges
        supabase.auth.admin.update_user_by_id(
            user_id,
            attributes={"password": new_password}
        )

        return jsonify({"message": f"Password for user {user_id} has been reset successfully."})
    except Exception as e:
        print("Error resetting password:", str(e))
        return jsonify({"error": str(e)}), 500

@app.route("/admin/register-pic", methods=["POST"])
def admin_register_pic():
    try:
        admin_user, auth_error = require_admin_request_user()
        if auth_error:
            return auth_error

        data = request.get_json() or {}
        name = str(data.get("name") or "").strip()
        username = str(data.get("username") or "").strip()
        email = str(data.get("email") or "").strip().lower()
        password = str(data.get("password") or "")
        role = str(data.get("role") or "user").strip().lower()
        valid_until_input = str(data.get("validUntil") or "").strip()

        if not name or not username or not email or not password:
            return jsonify({"error": "name, username, email, and password are required"}), 400

        if role not in {"admin", "user", "viewer"}:
            return jsonify({"error": "Invalid role supplied"}), 400

        existing_username = first_supabase_row(
            supabase.table("profiles").select("id").eq("username", username).limit(1).execute()
        )
        if existing_username:
            return jsonify({"error": "Username already exists"}), 409

        existing_email = first_supabase_row(
            supabase.table("profiles").select("id").eq("email", email).limit(1).execute()
        )
        if existing_email:
            return jsonify({"error": "Email already exists"}), 409

        valid_until = None
        if valid_until_input:
            try:
                valid_until = license_date_to_valid_until(valid_until_input)
            except ValueError:
                return jsonify({"error": "License date must use YYYY-MM-DD format"}), 400

        auth_response = supabase.auth.admin.create_user(
            {
                "email": email,
                "password": password,
                "email_confirm": True,
                "user_metadata": {
                    "must_change_password": True,
                    "temporary_password": True,
                    "temporary_password_set_at": datetime.utcnow().isoformat() + "Z",
                    "display_name": name,
                    "username": username,
                    "created_by_admin_id": admin_user["id"],
                },
            }
        )

        created_user = auth_response_to_user(auth_response)
        uid = get_auth_user_field(created_user, "id")
        if not uid:
            raise RuntimeError("Failed to get new user ID after auth registration.")

        supabase.table("profiles").upsert(
            {
                "id": uid,
                "display_name": name,
                "email": email,
                "username": username,
            },
            on_conflict="id",
        ).execute()

        supabase.table("user_roles").upsert(
            {
                "id": uid,
                "role": role,
            },
            on_conflict="id",
        ).execute()

        if valid_until:
            supabase.table("licenses").upsert(
                {
                    "id": uid,
                    "valid_until": valid_until,
                },
                on_conflict="id",
            ).execute()

        return jsonify({
            "message": "User successfully registered without sending a signup email.",
            "user_id": uid,
        })
    except Exception as exc:
        app.logger.exception("Failed to register PIC")
        return jsonify({"error": str(exc)}), 500




@app.route('/upload_FCSHPC_OSDP', methods=['POST'])
def upload_files():
    if 'files' not in request.files:
        return jsonify({"error": "No files provided"}), 400

    files = request.files.getlist('files')
    if not files or len(files) == 0:
        return jsonify({"error": "Empty file list"}), 400

    merged_df = pd.DataFrame()

    for file in files:
        if file.filename == '':
            continue
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)
        try:
            # Read specific columns from Excel (adjust column indices as needed)
            df = pd.read_excel(filepath, skiprows=2, usecols=[8, 9, 13, 16, 17, 18])
            merged_df = pd.concat([merged_df, df], ignore_index=True)
        except Exception as e:
            return jsonify({"error": f"Error reading {file.filename}: {str(e)}"}), 500

    # Clean and prepare data
    merged_df['Distributor'] = merged_df['Distributor'].ffill()
    merged_df['Distributor Name'] = merged_df['Distributor Name'].ffill()
    
    # Get sorting parameters
    primary_sort = request.args.get('primary_sort', 'Distributor')
    secondary_sort = request.args.get('secondary_sort', 'Sales Route')
    primary_asc = request.args.get('primary_asc', 'true').lower() == 'true'
    secondary_asc = request.args.get('secondary_asc', 'true').lower() == 'true'

    # Sort the data
    sorted_df = merged_df.sort_values(
        by=[primary_sort, secondary_sort],
        ascending=[primary_asc, secondary_asc]
    )

     # Create simplified summary - just distributor code and count
    summary_df = sorted_df.groupby(['Distributor','Distributor Name']).size().reset_index(name='Total Data')

    return jsonify({
        "sorted_data": sorted_df.to_dict(orient='records'),
        "summary_data": summary_df.to_dict(orient='records')
    })

@app.route('/upload_FCSHPC_PBI', methods=['POST'])
def upload_files_pbi():
    if 'files1' not in request.files:
        return jsonify({"error": "No files provided"}), 400

    files = request.files.getlist('files1')
    if not files or len(files) == 0:
        return jsonify({"error": "Empty file list"}), 400

    merged_df1 = pd.DataFrame()

    for file in files:
        if file.filename == '':
            continue
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)
        try:
            # Read specific columns (adjust as per your actual column names or indices)
            df1 = pd.read_excel(filepath, usecols=[8,9,13,18,22,26])
            df1.drop(df1.tail(3).index,inplace=True)
            merged_df1 = pd.concat([merged_df1, df1], ignore_index=True)
        except Exception as e:
            return jsonify({"error": f"Error reading {file.filename}: {str(e)}"}), 500
    

    # Sorting
    primary_sort = request.args.get('primary_sort', 'Distributor')
    secondary_sort = request.args.get('secondary_sort', 'Sales Route')
    primary_asc = request.args.get('primary_asc', 'true').lower() == 'true'
    secondary_asc = request.args.get('secondary_asc', 'true').lower() == 'true'

    sorted_df = merged_df1.sort_values(
        by=[primary_sort, secondary_sort],
        ascending=[primary_asc, secondary_asc]
    )

    # Summary
    summary_df = sorted_df.groupby(['Distributor', 'Distributor Name']).size().reset_index(name='Total Data')

    return jsonify({
        "sorted_data_PBI": sorted_df.to_dict(orient='records'),
        "summary_data_PBI": summary_df.to_dict(orient='records')
    })

@app.route('/clear', methods=['POST'])
def clear_data():
    try:
        for filename in os.listdir(UPLOAD_FOLDER):
            file_path = os.path.join(UPLOAD_FOLDER, filename)
            if os.path.isfile(file_path):
                os.remove(file_path)
        return jsonify({"message": "All uploaded files deleted."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
@app.route('/clear_pbi', methods=['POST'])
def clear_pbi_data():
    files = glob.glob(os.path.join(UPLOAD_FOLDER, '*'))
    for f in files:
        os.remove(f)
    return jsonify({"message": "Power BI data cleared and files deleted"}), 200


#-------------------------reconcile process-------------------------------------

@app.route('/reconcile_all', methods=['POST'])
def reconcile_all_data():
    import time
    import uuid
    import os
    import pickle
    import numpy as np
    start_time = time.time()
    try:
        data = request.get_json()
        osdp_data = data.get('osdp_data', [])
        pbi_data = data.get('pbi_data', [])
        business_type = data.get('businessType', '').strip()
        report_type = data.get('reportType', '').strip()

        osdp_df = pd.DataFrame(osdp_data)
        pbi_df = pd.DataFrame(pbi_data)
        osdp_df.columns = [c.strip() for c in osdp_df.columns]
        pbi_df.columns = [c.strip() for c in pbi_df.columns]
        
        report_type_name = data.get('reportTypeName', '').strip()
        is_dss = 'Daily Sales Summary' in report_type or 'Daily Sales Summary' in report_type_name

        dss_report_rows = []
        # --- DSS AGGREGATION & CARTESIAN PRODUCT PREVENTION ---
        if is_dss:
            # 1. First, build the exact DSS rows we need to render the DSS Table UI
            dss_report_rows = build_dss_aggregated_rows(osdp_df, pbi_df)
            
            # 2. To avoid OOM/Cartesian exploding the server on the generic row-by-row tracking:
            # Reassign osdp_df and pbi_df to completely aggregated single-row-per-distributor versions.
            # This causes the generic engine below to diff 1-to-1 aggregated rows!
            osdp_df = build_daily_sales_export_frame(osdp_df)
            pbi_df = build_daily_sales_export_frame(pbi_df)

        # KEY COLUMNS
        if is_dss:
            key_cols = ['Distributor']
        elif business_type == 'HPC' and report_type in ['FCS HPC']:
            key_cols = ['Distributor', 'Sales Route']
        elif business_type == 'HPC' and report_type == 'EFOS Outlet':
            key_cols = ['Distributor', 'Sales Route', 'Outlet Code', 'Date']
        elif business_type == 'IC' and report_type == 'EFOS Outlet':
            key_cols = ['Distributor', 'Sales Route', 'Outlet Code', 'Date']
        elif business_type == 'HPC' and report_type == 'IQ Performance Outlet':
            key_cols = ['Distributor', 'Outlet Code']
        elif business_type == 'IC' and report_type == 'IC IQ Performance':
            key_cols = ['Distributor', 'Sales Route', 'Outlet Code']
        elif report_type == 'Raw Data Invoice Level':
            key_cols = ['Distributor', 'Outlet Code', 'Invoice No','Invoice Date','Prod Code']
        else:
            key_cols = ['Distributor', 'Sales Route']  # default

        # Column presence check
        missing_osdp_cols = [col for col in key_cols if col not in osdp_df.columns]
        missing_pbi_cols = [col for col in key_cols if col not in pbi_df.columns]
        if missing_osdp_cols or missing_pbi_cols:
            return jsonify({
                "error": "Missing required columns",
                "osdp_missing": missing_osdp_cols,
                "pbi_missing": missing_pbi_cols
            }), 400

        # Normalize date columns
        date_cols = [col for col in key_cols if 'date' in col.lower()]
        for df in [osdp_df, pbi_df]:
            for col in date_cols:
                if col in df.columns:
                    df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%Y-%m-%d')
                    df[col] = df[col].replace('NaT', '')

        # Normalize key columns to uniform strings (strip trailing .0) before merging
        for col in key_cols:
            if col in osdp_df.columns:
                osdp_df[col] = osdp_df[col].astype(str).str.replace(r'\.0$', '', regex=True).str.strip().replace(['nan', 'NaN', 'None'], '')
            if col in pbi_df.columns:
                pbi_df[col] = pbi_df[col].astype(str).str.replace(r'\.0$', '', regex=True).str.strip().replace(['nan', 'NaN', 'None'], '')

        # Create key for fast merging
        osdp_df['key'] = osdp_df[key_cols].astype(str).agg(' - '.join, axis=1)
        pbi_df['key'] = pbi_df[key_cols].astype(str).agg(' - '.join, axis=1)

        print("OSDP columns:", osdp_df.columns.tolist())
        print("PBI columns:", pbi_df.columns.tolist())

        # --- Use set operations for unique/missing
        osdp_keys = set(osdp_df['key'])
        pbi_keys = set(pbi_df['key'])
        only_in_osdp = osdp_df[~osdp_df['key'].isin(pbi_keys)].copy()
        only_in_osdp['Mismatch Type'] = 'Missing in PBI'
        only_in_pbi = pbi_df[~pbi_df['key'].isin(osdp_keys)].copy()
        only_in_pbi['Mismatch Type'] = 'Missing in OSDP'

        # Exclude 'Outlet Sales' == 0 for Missing in OSDP when IC / IC IQ Performance
        if business_type == 'IC' and report_type == 'IC IQ Performance':
            if 'Sales Route' in only_in_osdp.columns:
                only_in_osdp = only_in_osdp[~only_in_osdp['Sales Route'].astype(str).str.startswith('SI')]
            if 'Sales Route' in only_in_pbi.columns:
                only_in_pbi = only_in_pbi[~only_in_pbi['Sales Route'].astype(str).str.startswith('SI')]

        # Ignore Missing in PBI for IC EFOS Outlet
        if business_type == 'IC' and report_type == 'EFOS Outlet':
            only_in_osdp = pd.DataFrame(columns=only_in_osdp.columns)

        # Merge for direct comparison (fast!)
        compare_columns = []
        for col in osdp_df.columns:
            if col in ['key', 'Distributor Name'] or col not in pbi_df.columns:
                continue
            
            # Exclude PJP Compliance and Geocode matched for IC EFOS Outlet
            if business_type == 'IC' and report_type == 'EFOS Outlet':
                if col.strip().lower() in ['pjp compliance', 'geocode matched', 'geo code matched']:
                    continue
            
            compare_columns.append(col)

        column_mapping_dict = {
            ('UFS', 'EFOS Outlet'): {
                'Time In':'Visit Start Time',     
                'Time Out':'Visit End Time',
                'Actual Outlet Time(in minutes)':'Avg Order Taking Time on Outlet',
                'Effective Outlet Time(in minutes)':'Effective Outlet Time',
                'Sales Value':'Sales Turnover',
                '#SKU':'SKU #',
            },
        }
        mapping_key = (business_type, report_type)
        column_mapping = column_mapping_dict.get(mapping_key, {})

        # Map columns in PBI for merged compare
        pbi_df_renamed = pbi_df.rename(
            columns={v: k for k, v in column_mapping.items()}
        )

        # Prepare for comparison
        osdp_comp = osdp_df.set_index('key')[compare_columns]
        pbi_comp = pbi_df_renamed.set_index('key')[compare_columns]
        merged = osdp_comp.join(pbi_comp, lsuffix='_osdp', rsuffix='_pbi', how='inner')

        # Pre-cache osdp rows by key for fast lookup
        if osdp_df['key'].duplicated().any():
            print("WARNING: Duplicate keys found in OSDP, only using first occurrence.")
            dupes = osdp_df[osdp_df['key'].duplicated(keep=False)]
            print("Duplicate rows:\n", dupes)
        osdp_dict = osdp_df.drop_duplicates('key').set_index('key').to_dict('index')


        # Vectorized difference detection
        osdp_cols = [f"{col}_osdp" for col in compare_columns]
        pbi_cols  = [f"{col}_pbi" for col in compare_columns]
        
        # Helper to safely convert to float where possible, otherwise keep as string
        def safe_float_convert(s):
            try:
                # convert to float and round to 4 decimals to avoid tiny precision issues
                return round(float(s), 4)
            except (ValueError, TypeError):
                return str(s).strip()

        # We'll build the mask manually by comparing Series to handle floats vs strings gracefully
        unequal_mask_list = []
        for o_col, p_col in zip(osdp_cols, pbi_cols):
            # Apply safe conversion
            o_series = merged[o_col].apply(safe_float_convert)
            p_series = merged[p_col].apply(safe_float_convert)
            
            # They are unequal if the converted values don't match, AND neither is an empty string
            # (An empty string compared to NaN can sometimes be tricky, but we handled NaNs earlier)
            mask = (o_series != p_series) & ~((merged[o_col].astype(str).str.strip() == '') & (merged[p_col].astype(str).str.strip() == ''))
            unequal_mask_list.append(mask)

        # Transpose to get a boolean array of shape (rows, cols)
        unequal_mask = np.column_stack(unequal_mask_list)

        mismatched_values = []
        rows, cols = np.where(unequal_mask)
        for i, j in zip(rows, cols):
            key = merged.index[i]
            col = compare_columns[j]
            diffs = {col: {'OSDP': merged.iloc[i][f"{col}_osdp"], 'PBI': merged.iloc[i][f"{col}_pbi"]}}
            osdp_row = osdp_dict[key]
            result = {"Mismatch Type": "Value mismatch", "Differences": diffs}
            for k in key_cols:
                result[k] = osdp_row.get(k, '')
            if 'Distributor Name' in osdp_row and 'Distributor Name' not in key_cols:
                result['Distributor Name'] = osdp_row['Distributor Name']
            mismatched_values.append(result)

        # Combine all mismatches
        reconciliation_result = []
        reconciliation_result.extend(only_in_osdp.to_dict(orient='records'))
        reconciliation_result.extend(only_in_pbi.to_dict(orient='records'))
        reconciliation_result.extend(mismatched_values)

        # --- Add this logging block ---
        missing_in_osdp_count = sum(1 for row in reconciliation_result if row.get('Mismatch Type') == 'Missing in OSDP')
        missing_in_pbi_count  = sum(1 for row in reconciliation_result if row.get('Mismatch Type') == 'Missing in PBI')
        value_mismatch_count  = sum(1 for row in reconciliation_result if row.get('Mismatch Type') == 'Value mismatch')

        print(f"[RECONCILE] Missing in OSDP: {missing_in_osdp_count}")
        print(f"[RECONCILE] Missing in PBI: {missing_in_pbi_count}")
        print(f"[RECONCILE] Value mismatches: {value_mismatch_count}")

        if (missing_in_osdp_count + missing_in_pbi_count + value_mismatch_count) == 0:
            print("[RECONCILE] All records matched perfectly! 🎉")
        else:
            print("[RECONCILE] Some mismatches were found.")


        # Mismatch set
        mismatch_distributors = { str(item['Distributor']).strip() for item in reconciliation_result if 'Distributor' in item }

        # Generate summaries (fast, vectorized)
        summary_osdp = osdp_df[['Distributor', 'Distributor Name']].drop_duplicates().dropna()
        summary_osdp['Status'] = summary_osdp['Distributor'].astype(str).str.strip().isin(mismatch_distributors)
        summary_osdp['Status'] = summary_osdp['Status'].map({True: 'Mismatch', False: 'Match'})
        summary_osdp = summary_osdp.to_dict(orient='records')

        summary_pbi = pbi_df[['Distributor', 'Distributor Name']].drop_duplicates().dropna()
        summary_pbi['Status'] = summary_pbi['Distributor'].astype(str).str.strip().isin(mismatch_distributors)
        summary_pbi['Status'] = summary_pbi['Status'].map({True: 'Mismatch', False: 'Match'})
        summary_pbi = summary_pbi.to_dict(orient='records')

        # (dss_report_rows is already populated at the top if DSS)
        def convert_np(obj):
            if isinstance(obj, (np.integer, np.int64)):
                return int(obj)
            if isinstance(obj, (np.floating, np.float64)):
                return float(obj)
            return str(obj)

        print("success recons in %.2f seconds" % (time.time() - start_time))
        result = {
            'summary_osdp': summary_osdp,
            'summary_pbi': summary_pbi,
            'reconciliation_result': reconciliation_result,
            'key_columns': key_cols,
            'dss_report_rows': dss_report_rows,
        }
        result_id = str(uuid.uuid4())
        RESULTS_DIR = "reconcile_results"
        os.makedirs(RESULTS_DIR, exist_ok=True)
        with open(os.path.join(RESULTS_DIR, f'{result_id}.pkl'), 'wb') as f:
            pickle.dump(result, f)
        print("success recons in %.2f seconds" % (time.time() - start_time))
        return jsonify({'result_id': result_id, 'key_columns': key_cols})

    except Exception as e:
        import traceback
        err_msg = f"Error in /reconcile_all: {str(e)}\n{traceback.format_exc()}"
        print(err_msg)
        with open("last_reconcile_error.txt", "w") as err_file:
            err_file.write(err_msg)
        return jsonify({"error": str(e)}), 500


@app.route('/get_reconcile_summary', methods=['GET'])
def get_reconcile_summary():
    result_id = request.args.get('result_id')
    file_path = os.path.join(RESULTS_DIR, f'{result_id}.pkl')
    print("[SUMMARY] Looking for file:", file_path)
    try:
        with open(file_path, 'rb') as f:
            result = pickle.load(f)
        print("[SUMMARY] Loaded keys:", result.keys())
        return jsonify({
            'summary_osdp': result['summary_osdp'],
            'summary_pbi': result['summary_pbi'],
            'key_columns': result['key_columns'],
            'dss_report_rows': result.get('dss_report_rows', []),
        })
    except Exception as e:
        print("[SUMMARY ERROR]", e)
        return jsonify({'error': str(e)}), 404


@app.route('/get_reconcile_page', methods=['GET'])
def get_reconcile_page():
    result_id = request.args.get('result_id')
    file_path = os.path.join(RESULTS_DIR, f'{result_id}.pkl')
    page = int(request.args.get('page', 1))
    size = int(request.args.get('size', 100))
    print("[PAGE] Looking for file:", file_path)
    try:
        with open(file_path, 'rb') as f:
            result = pickle.load(f)
        print("[PAGE] Loaded keys:", result.keys())
        data = result['reconciliation_result']
        print(f"[PAGE] Found {len(data)} records in reconciliation_result.")
        start = (page - 1) * size
        end = start + size
        paginated = data[start:end]

        def remove_nans(obj_item):
            import math
            import numpy as np
            if isinstance(obj_item, list):
                return [remove_nans(item) for item in obj_item]
            elif isinstance(obj_item, dict):
                return {k: remove_nans(v) for k, v in obj_item.items()}
            elif isinstance(obj_item, (float, np.floating)):
                if math.isnan(obj_item) or math.isinf(obj_item):
                    return None
                return float(obj_item)
            return obj_item

        paginated = remove_nans(paginated)

        def convert_np2(obj):
            import numpy as np
            if isinstance(obj, (np.integer, np.int64)):
                return int(obj)
            if isinstance(obj, (np.floating, np.float64)):
                return float(obj)
            return str(obj)

        return Response(
            json.dumps({
                'rows': paginated,
                'total': len(data),
                'page': page,
                'size': size
            }, default=convert_np2),
            mimetype='application/json'
        )
    except Exception as e:
        print("[PAGE ERROR]", e)
        return jsonify({'error': str(e)}), 404


# ─── DSS AGGREGATION HELPERS ─────────────────────────────────────────────────

DSS_COL_ALIASES = {
    'sales qty cs':    'Sales Qty CS',
    'sales_qty_cs':    'Sales Qty CS',
    'sale qty cs':     'Sales Qty CS',
    'qty cs':          'Sales Qty CS',
    'sale qty pc':     'Sale Qty PC',
    'sale_qty_pc':     'Sale Qty PC',
    'sales qty pc':    'Sale Qty PC',
    'qty pc':          'Sale Qty PC',
    'free total qty':  'Free Total Qty',
    'free_total_qty':  'Free Total Qty',
    'free qty':        'Free Total Qty',
    'gsv amount':      'GSV Amount',
    'gsv_amount':      'GSV Amount',
    'gsv(amount)':     'GSV Amount',
    'gsv':             'GSV Amount',
    'niv total':       'NIV Total',
    'niv_total':       'NIV Total',
    'niv(net invoice value)': 'NIV Total',
    'niv':             'NIV Total',
    'sales turn over': 'Sales Turn Over',
    'sales_turn_over': 'Sales Turn Over',
    'turnover':        'Sales Turn Over',
    'sales turnover':  'Sales Turn Over',
    'distributor':     'Distributor',
    'distributor name': 'Distributor Name',
}

DSS_NUMERIC_COLS = ['Sales Qty CS', 'Sale Qty PC', 'Free Total Qty', 'GSV Amount', 'NIV Total', 'Sales Turn Over']


def normalize_dss_columns(df):
    """Rename columns using DSS_COL_ALIASES (case-insensitive)."""
    rename_map = {}
    used = set()
    for col in df.columns:
        canonical = DSS_COL_ALIASES.get(str(col).strip().lower())
        if canonical and canonical not in used:
            rename_map[col] = canonical
            used.add(canonical)
    return df.rename(columns=rename_map)


def aggregate_dss_by_distributor(df, prefix):
    """Group invoice-level DSS rows by Distributor and sum numeric metrics."""
    df = normalize_dss_columns(df)
    if 'Distributor' not in df.columns:
        raise ValueError(f"'Distributor' column not found. Available: {df.columns.tolist()}")
    df['Distributor'] = df['Distributor'].ffill()
    if 'Distributor Name' in df.columns:
        df['Distributor Name'] = df['Distributor Name'].ffill()
    for col in DSS_NUMERIC_COLS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    present = [c for c in DSS_NUMERIC_COLS if c in df.columns]
    grp = ['Distributor'] + (['Distributor Name'] if 'Distributor Name' in df.columns else [])
    agg = df.groupby(grp, as_index=False)[present].sum()
    return agg.rename(columns={c: f'{prefix}_{c}' for c in present})


def build_dss_aggregated_rows(osdp_df, pbi_df):
    """Build one Supabase row per distributor: CSDP from OSDP file, OSDP col from PBI file."""
    try:
        csdp = aggregate_dss_by_distributor(osdp_df.copy(), 'CSDP')
        osdp_agg = aggregate_dss_by_distributor(pbi_df.copy(), 'OSDP')
        
        # Enforce string type for safe merging
        if 'Distributor' in csdp.columns:
            csdp['Distributor'] = csdp['Distributor'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
        if 'Distributor' in osdp_agg.columns:
            osdp_agg['Distributor'] = osdp_agg['Distributor'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
            
        merged = pd.merge(csdp, osdp_agg, on='Distributor', how='outer', suffixes=('', '_dup'))
    except Exception as e:
        print(f"[DSS BUILD] Aggregation/Merge error: {e}")
        return []
        
    if 'Distributor Name' not in merged.columns and 'Distributor Name_dup' in merged.columns:
        merged['Distributor Name'] = merged['Distributor Name_dup']
    merged = merged.drop(columns=[c for c in merged.columns if c.endswith('_dup')], errors='ignore')
    col_map = {
        'Distributor':           'distributor_code',
        'Distributor Name':      'distributor_name',
        'CSDP_Sales Qty CS':     'csdp_sales_qty_cs',
        'CSDP_Sale Qty PC':      'csdp_sale_qty_pc',
        'CSDP_Free Total Qty':   'csdp_free_total_qty',
        'CSDP_GSV Amount':       'csdp_gsv_amount',
        'CSDP_NIV Total':        'csdp_niv_total',
        'CSDP_Sales Turn Over':  'csdp_sales_turn_over',
        'OSDP_Sales Qty CS':     'osdp_sales_qty_cs',
        'OSDP_Sale Qty PC':      'osdp_sale_qty_pc',
        'OSDP_Free Total Qty':   'osdp_free_total_qty',
        'OSDP_GSV Amount':       'osdp_gsv_amount',
        'OSDP_NIV Total':        'osdp_niv_total',
        'OSDP_Sales Turn Over':  'osdp_sales_turn_over',
    }
    result = merged.rename(columns=col_map)
    present_cols = [c for c in col_map.values() if c in result.columns]
    result = result[present_cols]
    num_cols = [c for c in present_cols if c not in ('distributor_code', 'distributor_name')]
    result[num_cols] = result[num_cols].fillna(0)
    return result.to_dict(orient='records')

def build_hpc_dss_template_bytes(rows, year="", month="", business_type="HPC"):
    if not DSS_RECON_TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"DSS template not found: {DSS_RECON_TEMPLATE_PATH}")

    wb = load_workbook(DSS_RECON_TEMPLATE_PATH)
    sheet_name = "HPC Daily Sales Summary"
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Template sheet '{sheet_name}' not found.")

    ws = wb[sheet_name]
    for extra_sheet in list(wb.sheetnames):
        if extra_sheet != sheet_name:
            wb.remove(wb[extra_sheet])

    country_config = resolve_dss_template_country_config(business_type=business_type)
    master_distributors = get_active_master_distributors(
        business_type=business_type,
        allowed_countries=country_config["allowedCountries"],
    )
    if not master_distributors:
        allowed_display = ", ".join(country_config["allowedCountries"]) or "None"
        raise ValueError(
            f"No active {business_type} distributors found in master_distributors for allowed countries: {allowed_display}."
        )

    metric_map = {
        3: ("CSDP Sales Qty CS", "csdp_sales_qty_cs"),
        4: ("OSDP Sales Qty CS", "osdp_sales_qty_cs"),
        6: ("CSDP Sale Qty PC", "csdp_sale_qty_pc"),
        7: ("OSDP Sale Qty PC", "osdp_sale_qty_pc"),
        9: ("CSDP Free Total Qty", "csdp_free_total_qty"),
        10: ("OSDP Free Total Qty", "osdp_free_total_qty"),
        12: ("CSDP GSV Amount", "csdp_gsv_amount"),
        13: ("OSDP GSV Amount", "osdp_gsv_amount"),
        15: ("CSDP NIV Total", "csdp_niv_total"),
        16: ("OSDP NIV Total", "osdp_niv_total"),
        18: ("CSDP Sales Turn Over", "csdp_sales_turn_over"),
        19: ("OSDP Sales Turn Over", "osdp_sales_turn_over"),
    }
    formula_columns = {
        5: lambda row_idx: f"=C{row_idx}-D{row_idx}",
        8: lambda row_idx: f"=F{row_idx}-G{row_idx}",
        11: lambda row_idx: f"=I{row_idx}-J{row_idx}",
        14: lambda row_idx: f"=L{row_idx}-M{row_idx}",
        17: lambda row_idx: f"=O{row_idx}-P{row_idx}",
        20: lambda row_idx: f"=R{row_idx}-S{row_idx}",
    }

    def row_value(row, *keys):
        return pick_first_present(*[row.get(key) for key in keys])

    template_row_idx = 2
    template_styles = {
        col_idx: copy(ws.cell(row=template_row_idx, column=col_idx)._style)
        for col_idx in range(1, 22)
    }
    template_height = ws.row_dimensions[template_row_idx].height

    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    normalized_rows = {}
    for raw_row in rows or []:
        normalized = normalize_dss_export_row(raw_row)
        distributor_code = normalized.get("distributor_code", "")
        if distributor_code:
            normalized_rows[distributor_code] = normalized

    matched = 0
    for row_idx, master_row in enumerate(master_distributors, start=2):
        if template_height is not None:
            ws.row_dimensions[row_idx].height = template_height

        for col_idx in range(1, 22):
            ws.cell(row=row_idx, column=col_idx)._style = copy(template_styles[col_idx])

        distributor_code = master_row["code"]
        dss_row = normalized_rows.get(distributor_code, {})

        ws.cell(row=row_idx, column=1).value = distributor_code
        ws.cell(row=row_idx, column=2).value = (
            master_row["name"]
            or dss_row.get("distributor_name")
            or ""
        )

        for col_idx, keys in metric_map.items():
            value = safe_metric_number(row_value(dss_row, *keys))
            ws.cell(row=row_idx, column=col_idx).value = value

        for col_idx, formula_factory in formula_columns.items():
            ws.cell(row=row_idx, column=col_idx).value = formula_factory(row_idx)

        ws.cell(row=row_idx, column=21).value = ""
        if dss_row:
            matched += 1

    try:
        wb.calculation.forceFullCalc = True
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output, matched


# ─── DSS UPLOAD ENDPOINTS ─────────────────────────────────────────────────────

@app.route('/upload_HPCDSS_OSDP', methods=['POST'])
def upload_hpcdss_osdp():
    """Upload HPC Daily Sales Summary OSDP Excel file (invoice-level rows)."""
    if 'files' not in request.files:
        return jsonify({"error": "No files provided"}), 400
    files = request.files.getlist('files')
    merged_df = pd.DataFrame()
    for file in files:
        if not file.filename:
            continue
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)
        try:
            df = pd.read_excel(filepath)
            df.columns = [str(c).strip() for c in df.columns]
            
            # Normalize and cut off footers per file before concatenating
            df = normalize_dss_columns(df)
            if 'Distributor' in df.columns:
                cutoff_mask = df['Distributor'].astype(str).str.strip().str.contains('^(Grand\\s+)?Total\\s*$|^Applied filters', case=False, regex=True)
                if cutoff_mask.any():
                    cutoff_idx = cutoff_mask.idxmax()
                    df = df.iloc[:cutoff_idx]
                    
            merged_df = pd.concat([merged_df, df], ignore_index=True)
        except Exception as e:
            return jsonify({"error": f"Error reading {file.filename}: {str(e)}"}), 500
            
    if merged_df.empty:
        return jsonify({"error": "No data found in uploaded files"}), 400
        
    if 'Distributor' in merged_df.columns:
        merged_df['Distributor'] = merged_df['Distributor'].ffill()
        
    # Like IC DSS, drop any embedded sub-total rows where key columns contain "Total"
    cols_to_check = [c for c in ['Distributor', 'Sales Route', 'Outlet Code', 'Invoice Date', 'Invoice No'] if c in merged_df.columns]
    if cols_to_check:
        mask = ~merged_df[cols_to_check].apply(lambda row: row.astype(str).str.contains('Total', case=False, na=False)).any(axis=1)
        merged_df = merged_df[mask].reset_index(drop=True)
        
    if 'Distributor Name' in merged_df.columns:
        merged_df['Distributor Name'] = merged_df['Distributor Name'].ffill()
        
    for dt_col in merged_df.select_dtypes(include=['datetime', 'datetimetz']).columns:
        merged_df[dt_col] = merged_df[dt_col].astype(str).replace('NaT', '')
        
    for dt_col in merged_df.select_dtypes(include=['datetime', 'datetimetz']).columns:
        merged_df[dt_col] = merged_df[dt_col].dt.strftime('%Y-%m-%d %H:%M:%S').fillna('')
    merged_df = merged_df.dropna(how='all')
    sort_by = [c for c in ['Distributor', 'Invoice Date'] if c in merged_df.columns]
    sorted_df = merged_df.sort_values(by=sort_by) if sort_by else merged_df
    grp = [c for c in ['Distributor', 'Distributor Name'] if c in sorted_df.columns]
    summary_df = sorted_df.groupby(grp).size().reset_index(name='Total Records') if grp else pd.DataFrame()
    return jsonify({
        "sorted_data": sorted_df.fillna('').to_dict(orient='records'),
        "summary_data": summary_df.to_dict(orient='records'),
    })


@app.route('/upload_HPCDSS_PBI', methods=['POST'])
def upload_hpcdss_pbi():
    """Upload HPC Daily Sales Summary Power BI Excel file (invoice-level rows)."""
    if 'files1' not in request.files:
        return jsonify({"error": "No files provided"}), 400
    files = request.files.getlist('files1')
    merged_parts = []
    
    for file in files:
        if not file.filename:
            continue
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)
        try:
            # PBI files sometimes have 2 blank rows at top, sometimes none
            try:
                df = pd.read_excel(filepath, skiprows=2)
                df.columns = [str(c).strip() for c in df.columns]
                # Check for signs of validity (a known column mapped in normalize_dss_columns)
                if not any(c.lower() in ('distributor', 'sold-to pt', 'document number', 'invoice no', 'sold-to party') for c in df.columns):
                    raise ValueError("No valid columns found with skiprows=2")
            except Exception:
                df = pd.read_excel(filepath)
                df.columns = [str(c).strip() for c in df.columns]

            # Clean each file before concatenating so one file's footer does not
            # truncate the data coming from later uploaded files.
            df = normalize_dss_columns(df)
            if 'Distributor' not in df.columns:
                return jsonify({
                    "error": f"Missing 'Distributor' column in {file.filename}. Available columns: {df.columns.tolist()}"
                }), 400

            cutoff_mask = df['Distributor'].astype(str).str.strip().str.contains(
                '^(Grand\\s+)?Total\\s*$|^Applied filters',
                case=False,
                regex=True
            )
            if cutoff_mask.any():
                cutoff_idx = cutoff_mask.idxmax()
                df = df.iloc[:cutoff_idx]

            df = df.dropna(how='all').copy()
            df['Distributor'] = df['Distributor'].ffill()

            if 'Distributor Name' in df.columns:
                df['Distributor Name'] = df['Distributor Name'].ffill()

            cols_to_check = [c for c in ['Distributor', 'Sales Route', 'Outlet Code', 'Invoice Date', 'Invoice No'] if c in df.columns]
            if cols_to_check:
                mask = ~df[cols_to_check].apply(
                    lambda row: row.astype(str).str.contains('Total', case=False, na=False)
                ).any(axis=1)
                df = df[mask].reset_index(drop=True)

            df = df.dropna(how='all').reset_index(drop=True)
            if not df.empty:
                merged_parts.append(df)
        except Exception as e:
            return jsonify({"error": f"Error reading {file.filename}: {str(e)}"}), 500
            
    if not merged_parts:
        return jsonify({"error": "No data found in uploaded files"}), 400

    merged_df = pd.concat(merged_parts, ignore_index=True)
        
    try:
        merged_df = normalize_dss_columns(merged_df)
        if 'Distributor' not in merged_df.columns:
            return jsonify({"error": f"Upload successful but missing 'Distributor' column. Available columns: {merged_df.columns.tolist()}"}), 400

        merged_df['Distributor'] = merged_df['Distributor'].ffill()
        if 'Distributor Name' in merged_df.columns:
            merged_df['Distributor Name'] = merged_df['Distributor Name'].ffill()
             
        for dt_col in merged_df.select_dtypes(include=['datetime', 'datetimetz']).columns:
            merged_df[dt_col] = merged_df[dt_col].astype(str).replace('NaT', '')
            
        merged_df = merged_df.dropna(how='all')
        sort_by = [c for c in ['Distributor', 'Invoice Date'] if c in merged_df.columns]
        sorted_df = merged_df.sort_values(by=sort_by) if sort_by else merged_df
        grp = [c for c in ['Distributor', 'Distributor Name'] if c in sorted_df.columns]
        summary_df = sorted_df.groupby(grp).size().reset_index(name='Total Records') if grp else pd.DataFrame()
        
        return jsonify({
            "sorted_data_PBI": sorted_df.fillna('').to_dict(orient='records'),
            "summary_data_PBI": summary_df.to_dict(orient='records'),
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Processing error: {str(e)}"}), 500


@app.route('/upload_HPCEFOSSALES_OSDP', methods=['POST'])
def upload_hpcefossales_osdp():
    if 'files' not in request.files:
        return jsonify({"error": "No files provided"}), 400

    files = request.files.getlist('files')
    if not files:
        return jsonify({"error": "Empty file list"}), 400

    merged_df = pd.DataFrame()

    for file in files:
        if not file.filename:
            continue
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)
        try:
            df = pd.read_excel(
                filepath,
                skiprows=2,
                usecols=[0, 1, 2, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 22, 27],
            )
            df.drop(df.tail(2).index, inplace=True)
            merged_df = pd.concat([merged_df, df], ignore_index=True)
        except Exception as e:
            return jsonify({"error": f"Error reading {file.filename}: {str(e)}"}), 500

    merged_df['Distributor'] = merged_df['Distributor'].ffill()
    merged_df['Distributor Name'] = merged_df['Distributor Name'].ffill()

    columns_to_round = [
        '#SKU / Actual Calls',
        'Effective Outlet Time /Actual Calls',
        'PJP Compliance %',
        'Total Time Spent / Working Days',
        'Total Transit Time / Working Days',
        'Effective Outlet Time / Salesman',
        'Effective Day %',
    ]

    for col in columns_to_round:
        if col in merged_df.columns:
            merged_df[col] = np.round(merged_df[col], decimals=3)

    primary_sort = request.args.get('primary_sort', 'Distributor')
    secondary_sort = request.args.get('secondary_sort', 'Sales Route')
    primary_asc = request.args.get('primary_asc', 'true').lower() == 'true'
    secondary_asc = request.args.get('secondary_asc', 'true').lower() == 'true'

    sorted_df = merged_df.sort_values(
        by=[primary_sort, secondary_sort],
        ascending=[primary_asc, secondary_asc],
    )

    summary_df = sorted_df.groupby(['Distributor', 'Distributor Name']).size().reset_index(name='Total Data')

    return jsonify({
        "sorted_data": sorted_df.to_dict(orient='records'),
        "summary_data": summary_df.to_dict(orient='records'),
    })


@app.route('/upload_HPCEFOSSALES_PBI', methods=['POST'])
def upload_hpcefossales_pbi():
    if 'files1' not in request.files:
        return jsonify({"error": "No files provided"}), 400

    files = request.files.getlist('files1')
    if not files:
        return jsonify({"error": "Empty file list"}), 400

    merged_df = pd.DataFrame()

    for file in files:
        if not file.filename:
            continue
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)
        try:
            df = pd.read_excel(
                filepath,
                usecols=[0, 1, 2, 5, 7, 8, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 23, 28],
            )
            df.drop(df.tail(3).index, inplace=True)

            columns_to_round = [
                '#SKU / Actual Calls',
                'Effective Outlet Time /Actual Calls',
                'PJP Compliance %',
                'Total Time Spent / Working Days',
                'Total Transit Time / Working Days',
                'Effective Outlet Time / Salesman',
                'Effective Day %',
            ]

            for col in columns_to_round:
                if col in df.columns:
                    df[col] = np.trunc(df[col] * (10 ** 6)) / (10 ** 6)
                    df[col] = np.round(df[col], decimals=3)

            df = df.fillna(0)
            merged_df = pd.concat([merged_df, df], ignore_index=True)
        except Exception as e:
            return jsonify({"error": f"Error reading {file.filename}: {str(e)}"}), 500

    primary_sort = request.args.get('primary_sort', 'Distributor')
    secondary_sort = request.args.get('secondary_sort', 'Sales Route')
    primary_asc = request.args.get('primary_asc', 'true').lower() == 'true'
    secondary_asc = request.args.get('secondary_asc', 'true').lower() == 'true'

    sorted_df = merged_df.sort_values(
        by=[primary_sort, secondary_sort],
        ascending=[primary_asc, secondary_asc],
    )

    summary_df = sorted_df.groupby(['Distributor', 'Distributor Name']).size().reset_index(name='Total Data')

    return jsonify({
        "sorted_data_PBI": sorted_df.to_dict(orient='records'),
        "summary_data_PBI": summary_df.to_dict(orient='records'),
    })


@app.route('/upload_HPCIQSALES_OSDP', methods=['POST'])
def upload_hpciqsales_osdp():
    if 'files' not in request.files:
        return jsonify({"error": "No files provided"}), 400

    files = request.files.getlist('files')
    if not files:
        return jsonify({"error": "Empty file list"}), 400

    merged_df = pd.DataFrame()

    for file in files:
        if not file.filename:
            continue
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)
        try:
            df = pd.read_excel(filepath, skiprows=2, usecols=[0, 1, 2, 5, 6, 7, 10, 11, 12, 25, 26, 27, 35, 36, 37, 55])
            merged_df = pd.concat([merged_df, df], ignore_index=True)
        except Exception as e:
            return jsonify({"error": f"Error reading {file.filename}: {str(e)}"}), 500

    merged_df['Distributor'] = merged_df['Distributor'].ffill()
    merged_df['Distributor Name'] = merged_df['Distributor Name'].ffill()

    primary_sort = request.args.get('primary_sort', 'Distributor')
    secondary_sort = request.args.get('secondary_sort', 'Sales Route')
    primary_asc = request.args.get('primary_asc', 'true').lower() == 'true'
    secondary_asc = request.args.get('secondary_asc', 'true').lower() == 'true'

    sorted_df = merged_df.sort_values(
        by=[primary_sort, secondary_sort],
        ascending=[primary_asc, secondary_asc],
    )

    summary_df = sorted_df.groupby(['Distributor', 'Distributor Name']).size().reset_index(name='Total Data')

    return jsonify({
        "sorted_data": sorted_df.to_dict(orient='records'),
        "summary_data": summary_df.to_dict(orient='records'),
    })


@app.route('/upload_HPCIQSALES_PBI', methods=['POST'])
def upload_hpciqsales_pbi():
    if 'files1' not in request.files:
        return jsonify({"error": "No files provided"}), 400

    files = request.files.getlist('files1')
    if not files:
        return jsonify({"error": "Empty file list"}), 400

    merged_df = pd.DataFrame()

    for file in files:
        if not file.filename:
            continue
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)
        try:
            df = pd.read_excel(filepath, usecols=[0, 1, 2, 7, 8, 9, 12, 13, 14, 27, 28, 29, 37, 38, 39, 57])
            df.drop(df.tail(2).index, inplace=True)
            df = df.fillna(0)
            if 'Distributor' in df.columns:
                df['Distributor'] = df['Distributor'].astype(float)
            merged_df = pd.concat([merged_df, df], ignore_index=True)
        except Exception as e:
            return jsonify({"error": f"Error reading {file.filename}: {str(e)}"}), 500

    primary_sort = request.args.get('primary_sort', 'Distributor')
    secondary_sort = request.args.get('secondary_sort', 'Sales Route')
    primary_asc = request.args.get('primary_asc', 'true').lower() == 'true'
    secondary_asc = request.args.get('secondary_asc', 'true').lower() == 'true'

    sorted_df = merged_df.sort_values(
        by=[primary_sort, secondary_sort],
        ascending=[primary_asc, secondary_asc],
    )

    summary_df = sorted_df.groupby(['Distributor', 'Distributor Name']).size().reset_index(name='Total Data')

    return jsonify({
        "sorted_data_PBI": sorted_df.to_dict(orient='records'),
        "summary_data_PBI": summary_df.to_dict(orient='records'),
    })


@app.route('/upload_HPCIQOUTLET_OSDP', methods=['POST'])
def upload_hpciqoutlet_osdp():
    t0 = time.time()
    if 'files' not in request.files:
        return jsonify({"error": "No files provided"}), 400

    files = request.files.getlist('files')
    if not files:
        return jsonify({"error": "Empty file list"}), 400

    required_columns = [
        'Distributor',
        'Distributor Name',
        'Outlet Code',
        'Ever Billed Lines',
        'Everbilled Lines Replenished',
        '% Everbilled Replenishment',
        'Red Lines',
        'Redline Replenishment',
        '% Redline Replenishment',
        'WP Lines',
        'WP Replenishment',
        '% WP Replenishment',
        'MHSKUs Lines (CoC)',
        'MHSKUs  Replenishment (CoC)',
        '% MHSKUs Replenishment (CoC)',
        'Sales Turnover',
    ]

    df_list = []
    file_errors = []

    for file in files:
        if not file.filename:
            continue
        try:
            header_df = pd.read_excel(file.stream, skiprows=2, nrows=0, engine='openpyxl')
            missing = [col for col in required_columns if col not in header_df.columns]
            if missing:
                file_errors.append(f"{file.filename} missing columns: {', '.join(missing)}")
                continue

            file.stream.seek(0)
            df = pd.read_excel(
                file.stream,
                skiprows=2,
                usecols=required_columns,
                engine='openpyxl',
            )
            df_list.append(df)
        except Exception as e:
            file_errors.append(f"{file.filename} error: {str(e)}")

    if file_errors:
        return jsonify({"error": "Some files failed to process", "details": file_errors}), 400

    if not df_list:
        return jsonify({"error": "No valid files uploaded"}), 400

    merged_df = pd.concat(df_list, ignore_index=True)

    for col in ['Distributor', 'Distributor Name', 'Outlet Code']:
        if col in merged_df.columns:
            merged_df[col] = merged_df[col].ffill()
    merged_df = merged_df.fillna('')

    if 'Actual Outlet Time(in minutes)' in merged_df.columns:
        merged_df['Actual Outlet Time(in minutes)'] = pd.to_numeric(
            merged_df['Actual Outlet Time(in minutes)'],
            errors='coerce',
        ).fillna(0)

    if 'Sales Turnover' in merged_df.columns:
        merged_df['Sales Turnover'] = merged_df['Sales Turnover'].apply(lambda x: f"{x:.2f}")

    primary_sort = request.args.get('primary_sort', 'Distributor')
    secondary_sort = request.args.get('secondary_sort', 'Outlet Code')
    primary_asc = request.args.get('primary_asc', 'true').lower() == 'true'
    secondary_asc = request.args.get('secondary_asc', 'true').lower() == 'true'
    sort_cols = [c for c in [primary_sort, secondary_sort] if c in merged_df.columns]
    sorted_df = merged_df.sort_values(by=sort_cols, ascending=[primary_asc, secondary_asc][:len(sort_cols)])

    if all(col in sorted_df.columns for col in ['Distributor', 'Distributor Name']):
        summary_df = sorted_df.groupby(['Distributor', 'Distributor Name']).size().reset_index(name='Total Data')
    else:
        summary_df = pd.DataFrame()

    for col in sorted_df.select_dtypes(include=['datetime', 'datetimetz']).columns:
        sorted_df[col] = sorted_df[col].astype(str).replace('NaT', '')

    sorted_df = sorted_df.where(~sorted_df.isna(), '')
    print(f"OSDP upload processed in {time.time() - t0:.2f} seconds")

    return jsonify({
        "sorted_data": sorted_df.to_dict(orient='records'),
        "summary_data": summary_df.to_dict(orient='records'),
    })


@app.route('/upload_HPCIQOUTLET_PBI', methods=['POST'])
def upload_hpciqoutlet_pbi():
    t0 = time.time()
    if 'files1' not in request.files:
        return jsonify({"error": "No files provided"}), 400

    files = request.files.getlist('files1')
    if not files:
        return jsonify({"error": "Empty file list"}), 400

    def is_number(value):
        try:
            float(value)
            return True
        except Exception:
            return False

    df_list = []

    for file in files:
        if not file.filename:
            continue
        try:
            df = pd.read_excel(
                file.stream,
                usecols=[0, 1, 6, 11, 12, 13, 16, 17, 18, 31, 32, 33, 41, 42, 43, 61],
                engine='openpyxl',
            )
            df.columns = [str(c).strip() for c in df.columns]

            if 'Distributor' not in df.columns:
                return jsonify({"error": f"Missing 'Distributor' column in {file.filename}"}), 400

            df['Distributor'] = pd.to_numeric(df['Distributor'], errors='coerce').fillna(0)
            df = df[df['Distributor'] != 0]

            if not df.empty:
                last_a = df.iloc[-1, 0]
                last_b = df.iloc[-1, 1]
                if str(last_a).strip() != '' and str(last_b).strip() != '':
                    if is_number(last_a) and is_number(last_b):
                        diff = int(float(last_a) - float(last_b))
                        if 0 < diff < len(df):
                            df = df.iloc[:-diff]
                    else:
                        print(f"[INFO] Skipping row-trim for {file.filename}: last_A/B not numeric ({last_a}, {last_b})")

            if 'Sales Turnover' in df.columns:
                df['Sales Turnover'] = np.round(df['Sales Turnover'], decimals=6)
                df['Sales Turnover'] = df['Sales Turnover'].apply(lambda x: f"{x:.2f}")

            df_list.append(df)
        except Exception as e:
            print(f"[ERROR] Failed to process {file.filename}: {str(e)}")
            return jsonify({"error": f"Error reading {file.filename}: {str(e)}"}), 500

    merged_df = pd.concat(df_list, ignore_index=True) if df_list else pd.DataFrame()

    primary_sort = request.args.get('primary_sort', 'Distributor')
    secondary_sort = request.args.get('secondary_sort', 'Outlet Code')
    primary_asc = request.args.get('primary_asc', 'true').lower() == 'true'
    secondary_asc = request.args.get('secondary_asc', 'true').lower() == 'true'
    sort_cols = [c for c in [primary_sort, secondary_sort] if c in merged_df.columns]

    if not sort_cols:
        sorted_df = merged_df
    else:
        sorted_df = merged_df.sort_values(
            by=sort_cols,
            ascending=[primary_asc, secondary_asc][:len(sort_cols)],
        )

    if all(col in sorted_df.columns for col in ['Distributor', 'Distributor Name']):
        summary_df = sorted_df.groupby(['Distributor', 'Distributor Name']).size().reset_index(name='Total Data')
    else:
        summary_df = pd.DataFrame()

    for col in sorted_df.select_dtypes(include=['datetime', 'datetimetz']).columns:
        sorted_df[col] = sorted_df[col].astype(str).replace('NaT', '')

    sorted_df = sorted_df.where(~sorted_df.isna(), '')
    summary_df = summary_df.where(~summary_df.isna(), '')

    print(f"PBI upload processed in {time.time() - t0:.2f} seconds")
    return jsonify({
        "sorted_data_PBI": sorted_df.to_dict(orient='records'),
        "summary_data_PBI": summary_df.to_dict(orient='records'),
    })


def write_reconciliation_excel(records, mode, business_type, report_type, creator):
    import xlsxwriter
    from datetime import datetime, timedelta
    import os, io

    # Clean report type for logic
    report_type_clean = report_type.strip().lower()

    # Dynamic key columns
    if report_type_clean == 'efos outlet':
        key_cols = ['Distributor', 'Distributor Name', 'Sales Route', 'Outlet Code', 'Date']
    elif report_type_clean == 'daily sales summary':
        key_cols = ['Distributor', 'Distributor Name', 'Sales Route', 'Outlet Code', 'Invoice No', 'Invoice Date']
    elif report_type_clean == 'ic iq performance':
        key_cols = ['Distributor', 'Distributor Name', 'Sales Route', 'Outlet Code']
    elif report_type_clean == 'raw data invoice level':
        key_cols = ['Distributor', 'Distributor Name', 'Outlet Code', 'Invoice No', 'Invoice Date', 'Prod Code']
    else:
        key_cols = ['Distributor', 'Distributor Name', 'Sales Route']

    # Dynamic headers
    include_field_columns = mode == 'all' or any(row.get('Mismatch Type') == 'Value mismatch' for row in records)
    headers = key_cols + ['Mismatch Type']
    if include_field_columns:
        headers += ['Field', 'OSDP', 'PBI']

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet(f"{report_type} Summary")

    # Define styles
    header_format = workbook.add_format({
        'bold': True, 'font_color': 'white', 'bg_color': '#4F81BD',
        'border': 1, 'align': 'center'
    })
    match_format = workbook.add_format({'bg_color': '#C6EFCE', 'border': 1})
    mismatch_format = workbook.add_format({'bg_color': '#F4CCCC', 'border': 1})
    text_format = workbook.add_format({'border': 1})

    # Headers
    headers = ['Distributor', 'Distributor Name', 'Status']
    for col, h in enumerate(headers):
        worksheet.write(0, col, h, header_format)
        worksheet.set_column(col, col, 20)

    # Rows
    for row_idx, row in enumerate(records, start=1):
        worksheet.write(row_idx, 0, row['Distributor'], text_format)
        worksheet.write(row_idx, 1, row['Distributor Name'], text_format)
        status_format = match_format if row['Status'] == 'Match' else mismatch_format
        worksheet.write(row_idx, 2, row['Status'], status_format)

    workbook.close()
    output.seek(0)
    return send_file(output, download_name="summary_report.xlsx", as_attachment=True)

@app.route('/export_result_excel', methods=['POST'])
def export_result_excel():
    data = request.get_json()
    records = data.get('records', [])
    mode = data.get('mode', 'current')
    business_type = data.get('businessType', 'N/A')
    report_type = data.get('reportType', 'N/A')
    creator = data.get('creator', 'Auto Generated')

    # Clean report type for logic
    report_type_clean = report_type.strip().lower()

    # Dynamic key columns
    if business_type == 'HPC' and report_type_clean == 'efos outlet':
        key_cols = ['Distributor', 'Distributor Name', 'Sales Route', 'Outlet Code', 'Date']
    elif business_type == 'IC' and report_type_clean == 'efos outlet':
        key_cols = ['Distributor', 'Distributor Name', 'Sales Route', 'Outlet Code', 'Date']
    elif business_type == 'HPC' and report_type_clean == 'iq performance outlet':
        key_cols = ['Distributor', 'Distributor Name', 'Outlet Code']
    elif business_type == 'HPC' and report_type_clean == 'daily sales summary':
        key_cols = ['Distributor', 'Distributor Name', 'Sales Route', 'Outlet Code', 'Invoice No','Invoice Date']
    elif business_type == 'IC' and report_type_clean == 'daily sales summary':
        key_cols = ['Distributor', 'Distributor Name', 'Sales Route', 'Outlet Code', 'Invoice No','Invoice Date']
    elif business_type == 'IC' and report_type_clean == 'ic iq performance':
        key_cols = ['Distributor', 'Distributor Name', 'Sales Route', 'Outlet Code']
    elif report_type_clean == 'raw data invoice level':
        key_cols = ['Distributor', 'Distributor Name', 'Outlet Code', 'Invoice No','Invoice Date','Prod Code']
    else:
        key_cols = ['Distributor', 'Distributor Name', 'Sales Route']

    # Dynamic headers
    include_field_columns = mode == 'all' or any(row.get('Mismatch Type') == 'Value mismatch' for row in records)
    headers = key_cols + ['Mismatch Type']
    if include_field_columns:
        headers += ['Field', 'OSDP', 'PBI']

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet('Reconciliation')

    # Excel formatting
    title_format = workbook.add_format({
        'bold': True, 'font_size': 16, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#DDEBF7', 'font_color': '#1F4E78'
    })
    subtitle_format = workbook.add_format({
        'italic': True, 'font_size': 11, 'align': 'left', 'valign': 'vcenter', 'bg_color': '#F2F2F2'
    })
    meta_format = workbook.add_format({'bg_color': '#F2F2F2'})
    header_format = workbook.add_format({
        'bold': True, 'text_wrap': True, 'valign': 'middle', 'align': 'center', 'bg_color': '#B4C6E7', 'border': 1
    })
    center_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1})
    left_format = workbook.add_format({'align': 'left', 'valign': 'vcenter', 'border': 1})
    value_mismatch_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'bg_color': '#F8CBAD'})
    missing_osdp_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'bg_color': '#FFE699'})
    missing_pbi_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'bg_color': '#BDD7EE'})
    highlight_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'bg_color': '#FFC7CE'})
    subtitle_value_format = workbook.add_format({
        'italic': True, 'font_size': 11, 'align': 'left', 'valign': 'vcenter', 'bg_color': '#F2F2F2'
    })

    # Time
    is_render = os.getenv("RENDER", "").lower() == "true"
    current_time = datetime.utcnow() + timedelta(hours=8) if is_render else datetime.now()
    formatted_time = current_time.strftime('%Y-%m-%d %H:%M')

    # --- DYNAMIC HEADER ALIGNMENT ---
    if report_type_clean == 'daily sales summary':
        label_col = 4  # Move 2 cells right (col F, index 5)
        merge_start = 5
        merge_end = 9  # F2:I2, F3:I3
        end_mergeA=3
    elif report_type_clean== 'efos outlet':
        label_col = 3 + 1  # Move 1 cell right (col E, index 4)
        merge_start = 5
        merge_end = 8  # E2:G2, E3:G3
        end_mergeA=3
    elif report_type_clean== 'ic iq performance':
        label_col = 3   # Move 1 cell right (col E, index 4)
        merge_start = 4
        merge_end = 7  # E2:G2, E3:G3
        end_mergeA=2
    elif report_type_clean == 'raw data invoice level':
        label_col = 4  # Move 2 cells right (col F, index 5)
        merge_start = 5
        merge_end = 9  # F2:I2, F3:I3
        end_mergeA=3
    else:
        label_col = 3      # Default (col D, index 3)
        merge_start = 4
        merge_end = 6  # E2:G2, E3:G3
        end_mergeA=2

    # Title row (merge full header range)
    worksheet.merge_range(0, 0, 0, len(headers) - 1, 'Mismatch Result Report', title_format)

    # "Created by"/"Created on" -- always merge B2:D2 and B3:D3
    worksheet.write(1, 0, 'Created by:', subtitle_format)
    worksheet.write(2, 0, 'Created on:', subtitle_format)
    worksheet.merge_range(1, 1, 1, end_mergeA, creator, subtitle_value_format)   # B2:D2
    worksheet.merge_range(2, 1, 2, end_mergeA, formatted_time, subtitle_value_format) # B3:D3

    # Business Type/Report Type
    worksheet.write(1, label_col, 'Business Type:', subtitle_format)
    worksheet.write(2, label_col, 'Report Type:', subtitle_format)
    worksheet.merge_range(1, merge_start, 1, merge_end, business_type, subtitle_value_format)
    worksheet.merge_range(2, merge_start, 2, merge_end, report_type, subtitle_value_format)

    # Meta row
    worksheet.merge_range(3, 0, 3, len(headers) - 1, '', meta_format)

    # Column headers
    for col, header in enumerate(headers):
        worksheet.write(4, col, header, header_format)

    col_widths = [len(header) for header in headers]
    row_idx = 5

    # Helper to get key fields for row
    def get_row_keys(row):
        return [row.get(col, '') for col in key_cols]

    for row in records:
        mismatch_type = row.get('Mismatch Type', '')
        row_keys = get_row_keys(row)
        # 1. Value mismatch, with Differences (field-level detail)
        if mismatch_type == 'Value mismatch' and 'Differences' in row:
            for field, values in row['Differences'].items():
                values_to_write = row_keys + [mismatch_type, field, values.get('OSDP', ''), values.get('PBI', '')]
                for col, val in enumerate(values_to_write):
                    fmt = (value_mismatch_format if headers[col] == 'Mismatch Type' else
                           highlight_format if headers[col] in ['OSDP','PBI'] and values.get('OSDP') != values.get('PBI') else
                           center_format if headers[col] in key_cols + ['OSDP','PBI'] else
                           left_format)
                    worksheet.write(row_idx, col, '-' if val in [None, ''] else val, fmt)
                    col_widths[col] = max(col_widths[col], len(str(val)))
                row_idx += 1
        # 2. Value mismatch, but no field details (should be rare)
        elif mismatch_type == 'Value mismatch':
            values_to_write = row_keys + [mismatch_type]
            if 'Field' in headers:
                values_to_write += ['', '', '']
            for col, val in enumerate(values_to_write):
                fmt = (value_mismatch_format if headers[col] == 'Mismatch Type' else
                       center_format if headers[col] in key_cols + ['OSDP','PBI'] else
                       left_format)
                worksheet.write(row_idx, col, '-' if val in [None, ''] else val, fmt)
                col_widths[col] = max(col_widths[col], len(str(val)))
            row_idx += 1
        # 3. All mismatches (mode == 'all')
        elif mode == 'all':
            values_to_write = row_keys + [mismatch_type]
            if 'Field' in headers:
                values_to_write += ['', '', '']
            for col, val in enumerate(values_to_write):
                fmt = (missing_osdp_format if headers[col] == 'Mismatch Type' and mismatch_type == 'Missing in OSDP' else
                       missing_pbi_format if headers[col] == 'Mismatch Type' and mismatch_type == 'Missing in PBI' else
                       center_format if headers[col] in key_cols + ['OSDP','PBI'] else
                       left_format)
                worksheet.write(row_idx, col, '-' if val in [None, ''] else val, fmt)
                col_widths[col] = max(col_widths[col], len(str(val)))
            row_idx += 1
        # 4. Other cases
        else:
            values_to_write = row_keys + [mismatch_type]
            for col, val in enumerate(values_to_write):
                fmt = (missing_osdp_format if headers[col] == 'Mismatch Type' and mismatch_type == 'Missing in OSDP' else
                       missing_pbi_format if headers[col] == 'Mismatch Type' and mismatch_type == 'Missing in PBI' else
                       center_format if headers[col] in key_cols else
                       left_format)
                worksheet.write(row_idx, col, '-' if val in [None, ''] else val, fmt)
                col_widths[col] = max(col_widths[col], len(str(val)))
            row_idx += 1

    for col, width in enumerate(col_widths):
        if include_field_columns and headers[col] in ['OSDP','PBI']:
            worksheet.set_column(col, col, 12)
        else:
            worksheet.set_column(col, col, width + 2)

    worksheet.autofilter(4, 0, 4, len(headers) - 1)
    worksheet.freeze_panes(5, 0)
    worksheet.hide_gridlines(2)

    workbook.close()
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=f'Reconciliation_{mode}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# 🔧 Utility: Smart date parser to detect US vs EU formats
def smart_parse_date(series):
    parsed_us = pd.to_datetime(series, errors='coerce', dayfirst=False)
    parsed_eu = pd.to_datetime(series, errors='coerce', dayfirst=True)

    us_valid = parsed_us.notna().sum()
    eu_valid = parsed_eu.notna().sum()

    return parsed_eu if eu_valid > us_valid else parsed_us

# 📌 Route 1: Get columns for user selection
@app.route('/get_columns', methods=['POST'])
def get_columns():
    file = request.files['file']
    ext = file.filename.split('.')[-1]

    if ext == 'csv':
        df = pd.read_csv(file, nrows=0)
    else:
        df = pd.read_excel(file, nrows=0)

    return {'columns': df.columns.tolist()}


# 📌 Route 2: Convert selected date columns to user-specified format
@app.route('/convert_date', methods=['POST'])
def convert_date():
    file = request.files['file']
    columns = request.form.get('columns')
    date_format = request.form.get('format') or 'DD/MM/YYYY'

    # Convert custom format to Python strftime
    format_map = {
        'DD/MM/YYYY': '%d/%m/%Y',
        'DD/MM/YYYY HH:mm:ss': '%d/%m/%Y %H:%M:%S',
        'YYYY-MM-DD': '%Y-%m-%d',
    }
    strf_format = format_map.get(date_format, '%d/%m/%Y')

    # Load file
    if file.filename.endswith('.csv'):
        df = pd.read_csv(file)
    else:
        df = pd.read_excel(file)

    # Load selected columns from frontend
    try:
        cols = list(eval(columns)) if columns else []
    except Exception:
        cols = []

    # Convert each selected column
    for col in cols:
        if col in df.columns:
            try:
                parsed = smart_parse_date(df[col])
                df[col] = parsed.dt.strftime(strf_format)
            except Exception as e:
                print(f"Failed to parse column {col}: {e}")

    # Return converted Excel file
    output = io.BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)
    return send_file(output, download_name='converted_dates.xlsx', as_attachment=True)

scope = [
    'https://spreadsheets.google.com/feeds',
    'https://www.googleapis.com/auth/drive'
]

worksheet = None  # will be set by init_sheets()

def init_sheets(creds_arg: str | None = None):
    """Initializes global 'worksheet' once, using CLI arg/env/exe-dir."""
    global worksheet
    if worksheet is not None:
        return  # already initialized

    creds_json_env = os.environ.get("GOOGLE_CREDS")
    if creds_json_env:
        creds_dict = json.loads(creds_json_env)
        credentials = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
    else:
        cred_path = find_credentials(creds_arg)  # uses --creds, GOOGLE_APPLICATION_CREDENTIALS, EXE dir, etc.
        credentials = ServiceAccountCredentials.from_json_keyfile_name(cred_path, scope)

    gc = gspread.authorize(credentials)
    SHEET_ID = '1ql1BfkiuRuU3A3mfOxEw_GoL2gP5ki7eQECHxyfvFwk'
    # cache the worksheet handle
    worksheet = gc.open_by_key(SHEET_ID).worksheet('Summary')

@app.route("/export_to_sheets", methods=["POST"])
def export_to_sheets():
    data = request.get_json(silent=True) or {}

    year_raw = str(data.get("year", "")).strip()
    month = str(data.get("month", "")).strip()
    business_type = str(data.get("businessType", "")).strip()
    report_type = str(data.get("reportType", "")).strip()
    source = str(data.get("source", "")).strip()
    updated_by = (
        str(data.get("displayName", "")).strip()
        or str(data.get("updatedBy", "")).strip()
        or "Auto Generated"
    )
    records = data.get("records") or []
    report_type_id = str(data.get("reportTypeId", "")).strip() or report_type
    report_type_name = str(data.get("reportTypeName", "")).strip() or report_type

    if not year_raw or not month or not business_type or not report_type:
        return jsonify({"status": "error", "error": "Missing year/month/businessType/reportType"}), 400

    if not isinstance(records, list) or len(records) == 0:
        return jsonify({"status": "error", "error": "records[] is required"}), 400

    try:
        year = int(year_raw)
    except (TypeError, ValueError):
        return jsonify({"status": "error", "error": f"Invalid year '{year_raw}'"}), 400

    try:
        month_no = month_to_number(month)
        period_id = make_period_id(year, month)
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)}), 400

    updated = 0
    locked = 0
    errors = []

    for row in records:
        try:
            distributor_code = str(row.get("Distributor", "")).strip()
            distributor_name = str(row.get("Distributor Name", "")).strip()
            status = norm_status_ui(row.get("Status", ""))
            remark = str(row.get("Remark", "")).strip()

            if not distributor_code:
                errors.append({"row": row, "error": "Missing Distributor"})
                continue

            cell_id = make_cell_id(period_id, business_type, distributor_code, report_type_id)
            now_iso = datetime.utcnow().isoformat() + "Z"

            prev = first_supabase_row(
                supabase.table("recon_cells")
                .select("id,status,recons_no,created_at,created_by")
                .eq("id", cell_id)
                .limit(1)
                .execute()
            )
            legacy_prev = get_legacy_recon_cell(cell_id)

            prev_status = str((prev or {}).get("status", "")).strip().lower()
            legacy_status = str((legacy_prev or {}).get("status", "")).strip().lower()
            if prev_status == "match" or legacy_status == "match":
                locked += 1
                continue

            prev_recons = max(
                int((prev or {}).get("recons_no") or 0),
                int((legacy_prev or {}).get("reconsNo") or 0),
            )
            next_recons = prev_recons + 1
            created_at = (
                (prev or {}).get("created_at")
                or (legacy_prev or {}).get("createdAt")
                or now_iso
            )
            created_by = (
                (prev or {}).get("created_by")
                or (legacy_prev or {}).get("createdBy")
                or updated_by
            )

            payload = {
                "id": cell_id,
                "period_id": period_id,
                "year": year,
                "month": month_no,
                "business_type": business_type,
                "distributor_code": distributor_code,
                "distributor_name": distributor_name,
                "report_type_id": report_type_id,
                "report_type_name": report_type_name,
                "status": status,
                "remark": remark,
                "updated_by": updated_by,
                "recons_no": next_recons,
                "created_at": created_at,
                "created_by": created_by,
                "updated_at": now_iso,
            }

            if prev:
                supabase.table("recon_cells").update(
                    {k: v for k, v in payload.items() if k != "id"}
                ).eq("id", cell_id).execute()
            else:
                supabase.table("recon_cells").insert(payload).execute()

            mirror_legacy_recon_cell(cell_id, {
                "id": cell_id,
                "periodId": period_id,
                "year": year,
                "month": month_no,
                "businessType": business_type,
                "distributorCode": distributor_code,
                "distributorName": distributor_name,
                "reportTypeId": report_type_id,
                "reportTypeName": report_type_name,
                "status": status,
                "remark": remark,
                "updatedBy": updated_by,
                "reconsNo": next_recons,
                "createdAt": created_at,
                "createdBy": created_by,
                "updatedAt": now_iso,
            })

            attempt_snapshot = {
                "cellId": cell_id,
                "recons_no": next_recons,
                "period_id": period_id,
                "year": year,
                "month": month_no,
                "business_type": business_type,
                "distributor_code": distributor_code,
                "distributor_name": distributor_name,
                "report_type_id": report_type_id,
                "report_type_name": report_type_name,
                "status": status,
                "remark": remark,
                "updated_by": updated_by,
                "updated_at": now_iso,
                "client_updated_at": now_iso,
                "previous_status": (
                    (prev or {}).get("status")
                    or (legacy_prev or {}).get("status")
                    or ("no_data" if prev or legacy_prev else "new")
                ),
                "previous_recons_no": prev_recons,
                "source": source or "summary_export",
            }

            try:
                supabase.table("reconCells_attempts").insert(attempt_snapshot).execute()
            except Exception as attempt_exc:
                app.logger.warning("Skipping reconCells_attempts insert for %s: %s", cell_id, attempt_exc)

            updated += 1

        except Exception as e:
            errors.append({"row": row, "error": str(e)})

    status_code = 200
    if errors and updated == 0:
        status_code = 500
    elif errors:
        status_code = 207

    return jsonify({
        "status": "success",
        "periodId": period_id,
        "updated": updated,
        "locked": locked,
        "errors": errors
    }), status_code

@app.route("/export_dss_to_supabase", methods=["POST"])
def export_dss_to_supabase():
    data = request.get_json(silent=True) or {}

    result_id = str(data.get("resultId") or data.get("result_id") or "").strip()
    year_raw = str(data.get("year", "")).strip()
    month_raw = str(data.get("month", "")).strip()
    business_type = str(data.get("businessType", "")).strip() or "HPC"
    report_type_id = str(data.get("reportTypeId", "")).strip()
    report_type_name = str(data.get("reportTypeName", "")).strip() or "Daily Sales Summary"

    if not result_id:
        return jsonify({"status": "error", "error": "Missing result_id"}), 400
    if not year_raw or not month_raw:
        return jsonify({"status": "error", "error": "Missing year/month"}), 400

    try:
        year = int(year_raw)
    except (TypeError, ValueError):
        return jsonify({"status": "error", "error": f"Invalid year '{year_raw}'"}), 400

    try:
        month_no = month_to_number(month_raw)
        period_id = make_period_id(year, month_raw)
    except Exception as exc:
        return jsonify({"status": "error", "error": str(exc)}), 400

    file_path = os.path.join(RESULTS_DIR, f"{result_id}.pkl")
    if not os.path.exists(file_path):
        return jsonify({"status": "error", "error": f"Result '{result_id}' not found"}), 404

    try:
        with open(file_path, "rb") as handle:
            result = pickle.load(handle)
    except Exception as exc:
        return jsonify({"status": "error", "error": f"Failed to load result: {exc}"}), 500

    dss_rows = result.get("dss_report_rows") or []
    if not isinstance(dss_rows, list) or not dss_rows:
        return jsonify({
            "status": "error",
            "error": "No Daily Sales Summary rows found in this reconciliation result.",
        }), 400

    report_type_key = report_type_id or report_type_name
    now_iso = datetime.utcnow().isoformat() + "Z"
    payloads = []

    for raw_row in dss_rows:
        row = normalize_dss_export_row(raw_row)
        distributor_code = row.get("distributor_code")
        if not distributor_code:
            continue

        payloads.append({
            "id": f"{period_id}__{business_type}__{report_type_key}__{distributor_code}",
            "year": year,
            "month": month_no,
            "business_type": business_type,
            "distributor_code": distributor_code,
            "distributor_name": row.get("distributor_name", ""),
            "csdp_sales_qty_cs": row.get("csdp_sales_qty_cs", 0.0),
            "osdp_sales_qty_cs": row.get("osdp_sales_qty_cs", 0.0),
            "csdp_sale_qty_pc": row.get("csdp_sale_qty_pc", 0.0),
            "osdp_sale_qty_pc": row.get("osdp_sale_qty_pc", 0.0),
            "csdp_free_total_qty": row.get("csdp_free_total_qty", 0.0),
            "osdp_free_total_qty": row.get("osdp_free_total_qty", 0.0),
            "csdp_gsv_amount": row.get("csdp_gsv_amount", 0.0),
            "osdp_gsv_amount": row.get("osdp_gsv_amount", 0.0),
            "csdp_niv_total": row.get("csdp_niv_total", 0.0),
            "osdp_niv_total": row.get("osdp_niv_total", 0.0),
            "csdp_sales_turn_over": row.get("csdp_sales_turn_over", 0.0),
            "osdp_sales_turn_over": row.get("osdp_sales_turn_over", 0.0),
            "updated_at": now_iso,
        })

    if not payloads:
        return jsonify({"status": "error", "error": "No valid Daily Sales Summary rows to export."}), 400

    try:
        dropped_columns = upsert_payloads_with_schema_retry(
            DSS_REPORT_TABLE,
            payloads,
            on_conflict="id",
            chunk_size=500,
        )
    except Exception as exc:
        message = str(exc)
        if "does not exist" in message.lower():
            message = (
                f"Supabase table '{DSS_REPORT_TABLE}' is missing. "
                "Create it first before exporting the Daily Sales Summary report."
            )
        return jsonify({"status": "error", "error": message}), 500

    return jsonify({
        "status": "success",
        "table": DSS_REPORT_TABLE,
        "periodId": period_id,
        "exported": len(payloads),
        "skippedColumns": dropped_columns,
    }), 200

@app.route("/dss_template_country_config", methods=["GET", "POST"])
def dss_template_country_config():
    if request.method == "GET":
        business_type = str(request.args.get("businessType", "")).strip().upper() or "HPC"
        try:
            result = resolve_dss_template_country_config(business_type=business_type)
        except Exception as exc:
            return jsonify({"status": "error", "error": str(exc)}), 500
        return jsonify({"status": "success", **result}), 200

    data = request.get_json(silent=True) or {}
    business_type = str(data.get("businessType", "")).strip().upper() or "HPC"
    allowed_countries = data.get("allowedCountries")
    updated_by = str(data.get("updatedBy", "")).strip()

    if not isinstance(allowed_countries, list):
        return jsonify({"status": "error", "error": "allowedCountries must be a list."}), 400

    try:
        result = update_dss_template_country_config(
            business_type=business_type,
            allowed_countries=allowed_countries,
            updated_by=updated_by,
        )
    except Exception as exc:
        return jsonify({"status": "error", "error": str(exc)}), 500

    return jsonify({"status": "success", **result}), 200

@app.route("/export_dss_template", methods=["POST"])
def export_dss_template():
    data = request.get_json(silent=True) or {}
    business_type = str(data.get("businessType", "")).strip().upper() or "HPC"
    year = str(data.get("year", "")).strip()
    month = str(data.get("month", "")).strip()
    rows = data.get("rows") or []

    if business_type != "HPC":
        return jsonify({
            "status": "error",
            "error": "Template export currently supports the first sheet only: HPC Daily Sales Summary.",
        }), 400

    if not isinstance(rows, list):
        return jsonify({"status": "error", "error": "Invalid rows payload for template export."}), 400

    try:
        output, matched = build_hpc_dss_template_bytes(rows, year=year, month=month, business_type=business_type)
    except Exception as exc:
        return jsonify({"status": "error", "error": str(exc)}), 500

    month_slug = str(month or "All").replace(" ", "_")
    year_slug = str(year or "All").replace(" ", "_")
    filename = f"{business_type}_DSS_Recon_Template_{year_slug}_{month_slug}.xlsx"

    response = send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers["X-DSS-Matched-Rows"] = str(matched)
    return response

@app.route('/upload_FCSIC_OSDP', methods=['POST'])
def upload_files_FCSIC_OSDP():
    if 'files' not in request.files:
        return jsonify({"error": "No files provided"}), 400

    files = request.files.getlist('files')
    if not files:
        return jsonify({"error": "Empty file list"}), 400

    merged_df = pd.DataFrame()

    for file in files:
        if file.filename == '':
            continue

        try:
            file.stream.seek(0)
            df = read_excel_mixed(file, skiprows=2, usecols=[0, 1, 2, 4, 7, 8])
            merged_df = pd.concat([merged_df, df], ignore_index=True)
        except Exception as e:
            return jsonify({"error": f"Error reading {file.filename}: {str(e)}"}), 500


    # Clean and prepare data
    merged_df['Distributor'] = merged_df['Distributor'].ffill()
    merged_df['Distributor Name'] = merged_df['Distributor Name'].ffill()

    # Sorting
    primary_sort = request.args.get('primary_sort', 'Distributor')
    secondary_sort = request.args.get('secondary_sort', 'Sales Route')
    primary_asc = request.args.get('primary_asc', 'true').lower() == 'true'
    secondary_asc = request.args.get('secondary_asc', 'true').lower() == 'true'

    sorted_df = merged_df.sort_values(
        by=[primary_sort, secondary_sort],
        ascending=[primary_asc, secondary_asc]
    )

    summary_df = sorted_df.groupby(['Distributor','Distributor Name']).size().reset_index(name='Total Data')

    return jsonify({
        "sorted_data": sorted_df.to_dict(orient='records'),
        "summary_data": summary_df.to_dict(orient='records')
    })

@app.route('/upload_FCSIC_PBI', methods=['POST'])
def upload_files_FCSIC_pbi():
    if 'files1' not in request.files:
        return jsonify({"error": "No files provided"}), 400

    files = request.files.getlist('files1')
    if not files or len(files) == 0:
        return jsonify({"error": "Empty file list"}), 400

    merged_df1 = pd.DataFrame()

    for file in files:
        if file.filename == '':
            continue
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)
        try:
            # Read specific columns (adjust as per your actual column names or indices)
            df1 = pd.read_excel(filepath, usecols=[0,1,2,6,9,10])
            df1.drop(df1.tail(2).index,inplace=True)
            merged_df1 = pd.concat([merged_df1, df1], ignore_index=True)
        except Exception as e:
            return jsonify({"error": f"Error reading {file.filename}: {str(e)}"}), 500
    

    # Sorting
    primary_sort = request.args.get('primary_sort', 'Distributor')
    secondary_sort = request.args.get('secondary_sort', 'Sales Route')
    primary_asc = request.args.get('primary_asc', 'true').lower() == 'true'
    secondary_asc = request.args.get('secondary_asc', 'true').lower() == 'true'

    sorted_df = merged_df1.sort_values(
        by=[primary_sort, secondary_sort],
        ascending=[primary_asc, secondary_asc]
    )

    # Summary
    summary_df = sorted_df.groupby(['Distributor', 'Distributor Name']).size().reset_index(name='Total Data')

    return jsonify({
        "sorted_data_PBI": sorted_df.to_dict(orient='records'),
        "summary_data_PBI": summary_df.to_dict(orient='records')
    })

# ---------- JSON-SAFE HELPERS ----------
def _json_safe_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Make a dataframe safe to jsonify:
      - datetimes -> ISO strings 'YYYY-MM-DDTHH:MM:SS'
      - timedeltas -> total seconds (float)
      - NaN/NaT -> None
      - keep strings as-is; numpy scalars remain OK once converted via to_dict
    """
    if df is None or df.empty:
        return df

    out = df.copy()

    for col in out.columns:
        s = out[col]
        # Datetime-like -> ISO 8601 strings (drop tz for simplicity)
        if pd.api.types.is_datetime64_any_dtype(s) or pd.api.types.is_datetime64tz_dtype(s):
            s_dt = pd.to_datetime(s, errors='coerce')
            out[col] = s_dt.dt.strftime('%Y-%m-%dT%H:%M:%S')
            out.loc[s_dt.isna(), col] = None
        # Timedelta -> total seconds
        elif pd.api.types.is_timedelta64_dtype(s):
            out[col] = s.dt.total_seconds()
        else:
            # For other types, just ensure NaNs become None
            out[col] = s.where(~s.isna(), None)

    # Final NaN/NaT -> None sweep (needs object dtype to avoid None coercing back to NaN in float columns)
    out = out.astype(object).where(pd.notnull(out), None)
    return out


def _df_records(df: pd.DataFrame):
    """Return JSON-serializable records list."""
    if df is None or len(df) == 0:
        return []
    return _json_safe_df(df).to_dict(orient='records')


@app.route('/upload_ICEFOSSALES_OSDP', methods=['POST'])
def upload_files_IC_EFOS_sales_OSDP():
    # 1) Basic input validation
    if 'files' not in request.files:
        return jsonify({"error": "No files provided"}), 400

    files = request.files.getlist('files')
    if not files:
        return jsonify({"error": "Empty file list"}), 400

    merged_df = pd.DataFrame()

    REQUIRED_COLS = [
        'Distributor', 'Distributor Name', 'Sales Route',
        'Total Working Days', 'Planned Calls', 'Actual Calls',
        'Effective Outlet Time(in minutes)', 'Total Time Spent(in minutes)',
        'PJP Compliance', 'Time Spent', 'Geo Matched', 'Effective Day',
        'Effective Day %', 'Effective Outlet Time / Salesman',
        'Effective Outlet Time /Actual Calls', 'Total Time Spent / Working Days',
        'Total Transit Time / Working Days', 'PJP Compliance %',
        '#SKU / Actual Calls', '#SKU'
    ]

    # Your original indices:
    # USECOLS_IDX = [0,1,2,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,22,27]
    #
    # You asked to fill blanks in O,V,W,AA,AB => 14,21,22,26,27
    # So we MUST ensure those indices are included.
    USECOLS_IDX = [0, 1, 2, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 21, 22, 26, 27]

    SHEET_NAME = 'NGDMS CR EFOS Salesman View'

    # Excel-letter targets as indices (0-based): O=14, V=21, W=22, AA=26, AB=27
    FILL_ZERO_COL_IDXS = [21, 22, 26, 27]

    # 2) Read & merge all files (xls/xlsx mixed)
    for file in files:
        if not file or not file.filename:
            continue

        try:
            # IMPORTANT: read_excel_mixed consumes file stream via file.read()
            # So call it exactly once per upload file.
            try:
                df = read_excel_mixed(
                    file,
                    sheet_name=SHEET_NAME,
                    skiprows=2,
                    usecols=USECOLS_IDX
                )
            except Exception:
                df = read_excel_mixed(
                    file,
                    sheet_name=0,
                    skiprows=2,
                    usecols=USECOLS_IDX
                )

            # Drop footer rows if exist (2 total lines at bottom)
            if len(df) >= 2:
                df = df.iloc[:-2, :]

            # ✅ Fill blank cells in specific Excel columns to 0 BEFORE concat
            # This is done by position (works even if header names differ).
            for idx in FILL_ZERO_COL_IDXS:
                if 0 <= idx < df.shape[1]:
                    col_name = df.columns[idx]
                    df[col_name] = df[col_name].fillna(0)

            merged_df = pd.concat([merged_df, df], ignore_index=True)

        except Exception as e:
            return jsonify({"error": f"Error reading {file.filename}: {str(e)}"}), 500

    # 3) If nothing parsed
    if merged_df.empty:
        return jsonify({
            "sorted_data": [],
            "summary_data": [],
            "meta": {"rows": 0, "files": [f.filename for f in files if f and f.filename]}
        })

    # 4) Clean & type-fix
    for col in ['Distributor', 'Distributor Name', 'Sales Route']:
        if col in merged_df.columns:
            merged_df[col] = merged_df[col].astype('string').ffill()

    # Coerce obvious numeric columns early (robust against mixed inputs)
    numeric_like_cols = [
        '#SKU / Actual Calls',
        'Effective Outlet Time /Actual Calls',
        'PJP Compliance %',
        'Total Time Spent / Working Days',
        'Total Transit Time / Working Days',
        'Effective Outlet Time / Salesman',
        'Effective Day %',
        'Planned Calls', 'Actual Calls', 'Total Working Days',
        '#SKU'
    ]
    for c in numeric_like_cols:
        if c in merged_df.columns:
            merged_df[c] = pd.to_numeric(merged_df[c], errors='coerce')

    if '#SKU / Actual Calls' in merged_df.columns:
        merged_df['#SKU / Actual Calls'] = merged_df['#SKU / Actual Calls'].fillna(0)

    # Round selected columns (after coercion)
    columns_to_truncate = [
        '#SKU / Actual Calls',
        'Effective Outlet Time /Actual Calls',
        'PJP Compliance %',
        'Total Time Spent / Working Days',
        'Total Transit Time / Working Days',
        'Effective Outlet Time / Salesman',
        'Effective Day %'
    ]
    for col in columns_to_truncate:
        if col in merged_df.columns:
            merged_df[col] = merged_df[col].round(3)

    # 5) Sorting
    primary_sort = request.args.get('primary_sort', 'Distributor')
    secondary_sort = request.args.get('secondary_sort', 'Sales Route')
    primary_asc = request.args.get('primary_asc', 'true').lower() == 'true'
    secondary_asc = request.args.get('secondary_asc', 'true').lower() == 'true'

    sort_cols = [c for c in [primary_sort, secondary_sort] if c in merged_df.columns]
    if sort_cols:
        ascending_flags = [primary_asc] + ([secondary_asc] if len(sort_cols) > 1 else [])
        sorted_df = merged_df.sort_values(by=sort_cols, ascending=ascending_flags, kind='mergesort')
    else:
        sorted_df = merged_df

    # 6) Summary
    if {'Distributor', 'Distributor Name'}.issubset(sorted_df.columns):
        summary_df = (
            sorted_df
            .groupby(['Distributor', 'Distributor Name'], dropna=False)
            .size()
            .reset_index(name='Total Data')
        )
    else:
        summary_df = pd.DataFrame(columns=['Distributor', 'Distributor Name', 'Total Data'])

    # 7) JSON-safe return
    return jsonify({
        "sorted_data": _df_records(sorted_df),
        "summary_data": _df_records(summary_df),
        "meta": {
            "rows": int(len(sorted_df)),
            "groups": int(len(summary_df)),
            "columns": list(sorted_df.columns),
            "missing_required_cols": [c for c in REQUIRED_COLS if c not in sorted_df.columns]
        }
    })



@app.route('/upload_ICEFOSSALES_PBI', methods=['POST'])
def upload_files_IC_EFOS_Sales_pbi():
    if 'files1' not in request.files:
        return jsonify({"error": "No files provided"}), 400

    files = request.files.getlist('files1')
    if not files or len(files) == 0:
        return jsonify({"error": "Empty file list"}), 400

    merged_df1 = pd.DataFrame()

    for file in files:
        if file.filename == '':
            continue
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)
        try:
            # Read specific columns (adjust as per your actual column names or indices)
            df1 = pd.read_excel(filepath,usecols=[0,1,2,5,7,8,10,11,12,13,14,15,16,17,18,19,20,21,23,28])
            df1.drop(df1.tail(3).index,inplace=True)
            columns_to_truncate = ['#SKU / Actual Calls', 
                       'Effective Outlet Time /Actual Calls', 
                       'PJP Compliance %',
                       'Total Time Spent / Working Days',
                       'Total Transit Time / Working Days',
                       'Effective Outlet Time / Salesman',
                       'Effective Day %'
                       ]

            for col in columns_to_truncate:
                df1[col] = np.trunc(df1[col] * (10**6)) / (10**6)
                df1[col] = np.round(df1[col],decimals=3)

            df1 = df1.fillna(0)
            merged_df1 = pd.concat([merged_df1, df1], ignore_index=True)
        except Exception as e:
            return jsonify({"error": f"Error reading {file.filename}: {str(e)}"}), 500
    

    # Sorting
    primary_sort = request.args.get('primary_sort', 'Distributor')
    secondary_sort = request.args.get('secondary_sort', 'Sales Route')
    primary_asc = request.args.get('primary_asc', 'true').lower() == 'true'
    secondary_asc = request.args.get('secondary_asc', 'true').lower() == 'true'

    sorted_df = merged_df1.sort_values(
        by=[primary_sort, secondary_sort],
        ascending=[primary_asc, secondary_asc]
    )

    # Summary
    summary_df = sorted_df.groupby(['Distributor', 'Distributor Name']).size().reset_index(name='Total Data')

    return jsonify({
        "sorted_data_PBI": sorted_df.to_dict(orient='records'),
        "summary_data_PBI": summary_df.to_dict(orient='records')
    })

#--------------------------HPC EFOS Outlet----------------------------------

@app.route('/upload_HPCEFOSOUTLET_OSDP', methods=['POST'])
def upload_files_HPC_EFOS_Outlet_OSDP():
    import time
    t0 = time.time()
    if 'files' not in request.files:
        return jsonify({"error": "No files provided"}), 400

    files = request.files.getlist('files')
    if not files or len(files) == 0:
        return jsonify({"error": "Empty file list"}), 400

    # 1. Read all files into DataFrames (in memory)
    df_list = []
    for file in files:
        if file.filename == '':
            continue
        try:
            # Read directly from in-memory stream, not disk
            df = pd.read_excel(file.stream, usecols=[0,1,2,5,10,11,12,13,14,15,16], engine='openpyxl')
            df_list.append(df)
        except Exception as e:
            return jsonify({"error": f"Error reading {file.filename}: {str(e)}"}), 500

    # 2. Concatenate ONCE at the end
    if df_list:
        merged_df = pd.concat(df_list, ignore_index=True)
    else:
        merged_df = pd.DataFrame()

    # 3. Data Cleaning
    col_pk_fillin = ['Distributor', 'Distributor Name', 'Sales Route', 'Outlet Code']
    for col in col_pk_fillin:
        if col in merged_df.columns:
            merged_df[col] = merged_df[col].ffill()
    merged_df = merged_df.fillna('')
    if 'Actual Outlet Time(in minutes)' in merged_df.columns:
        merged_df['Actual Outlet Time(in minutes)'] = pd.to_numeric(
            merged_df['Actual Outlet Time(in minutes)'], errors='coerce'
        ).fillna(0)
    if 'Sales Value' in merged_df.columns:
        #merged_df['Sales Value'] = (merged_df['Sales Value'] * 100).astype(int) / 100
        merged_df['Sales Value'] = merged_df['Sales Value'].apply(lambda x: f"{x:.2f}")

    # 4. Sorting
    primary_sort = request.args.get('primary_sort', 'Distributor')
    secondary_sort = request.args.get('secondary_sort', 'Sales Route')
    primary_asc = request.args.get('primary_asc', 'true').lower() == 'true'
    secondary_asc = request.args.get('secondary_asc', 'true').lower() == 'true'
    sort_cols = [c for c in [primary_sort, secondary_sort] if c in merged_df.columns]
    sorted_df = merged_df.sort_values(by=sort_cols, ascending=[primary_asc, secondary_asc][:len(sort_cols)])

    # 5. Summary
    if all(col in sorted_df.columns for col in ['Distributor', 'Distributor Name']):
        summary_df = sorted_df.groupby(['Distributor', 'Distributor Name']).size().reset_index(name='Total Data')
    else:
        summary_df = pd.DataFrame()

    # 6. Fix datetime columns with NaT
    for col in sorted_df.select_dtypes(include=['datetime', 'datetimetz']).columns:
        sorted_df[col] = sorted_df[col].astype(str).replace('NaT', '')

    sorted_df = sorted_df.rename(
            columns={
            'Time In':'Visit Start Time',     
            'Time Out':'Visit End Time',
            'Actual Outlet Time(in minutes)':'Avg Order Taking Time on Outlet',
            'Effective Outlet Time(in minutes)':'Effective Outlet Time',
            'Sales Value':'Sales Turnover',
            '#SKU':'SKU #'}
        )

    sorted_df = sorted_df.where(~sorted_df.isna(), '')
    print(f"OSDP upload processed in {time.time() - t0:.2f} seconds")

    return jsonify({
        "sorted_data": sorted_df.to_dict(orient='records'),
        "summary_data": summary_df.to_dict(orient='records')
    })

@app.route('/upload_HPCEFOSOUTLET_PBI', methods=['POST'])
def upload_files_HPC_EFOS_Outlet_pbi():
    import time
    t0 = time.time()
    if 'files1' not in request.files:
        return jsonify({"error": "No files provided"}), 400

    files = request.files.getlist('files1')
    if not files or len(files) == 0:
        return jsonify({"error": "Empty file list"}), 400

    # 1. Read all files into DataFrames (in memory, no saving to disk)
    df_list = []
    for file in files:
        if file.filename == '':
            continue
        try:
            # Read directly from in-memory stream
            df1 = pd.read_excel(file.stream, usecols=[0, 1, 2, 5, 10, 11, 12, 13, 14, 15, 16], engine='openpyxl')
            # Drop last 3 rows (trailing summary rows)
            #if len(df1) >= 3:
            #    df1.drop(df1.tail(3).index, inplace=True)
            df1.columns = [c.strip() for c in df1.columns]
            df1.rename(columns={'# SKU': 'SKU #'}, inplace=True)
            df1 = df1.fillna('')

            # Convert numeric columns if present
            for col in ['SKU #', 'Sales Turnover', 'Avg Order Taking Time on Outlet']:
                if col in df1.columns:
                    df1[col] = pd.to_numeric(df1[col], errors='coerce').fillna(0)

            if 'Distributor' in df1.columns:
                df1['Distributor'] = pd.to_numeric(df1['Distributor'], errors='coerce').fillna(0)
            else:
                return jsonify({"error": f"Missing 'Distributor' column in {file.filename}"}), 400
            
            df1 = df1[df1['Distributor'] != 0]
            # --- Trim by last A/B for this file ---
            if not df1.empty:
                last_A = df1.iloc[-1, 0]
                last_B = df1.iloc[-1, 1]
                if str(last_A).strip() != '' and str(last_B).strip() != '':
                    try:
                        diff = int(float(last_A) - float(last_B))
                        if diff > 0 and diff < len(df1):
                            df1 = df1.iloc[:-diff]
                    except Exception as e:
                        print(f"[WARNING] Couldn't process row-trim by last A/B for {file.filename}: {e}")

            if 'Sales Turnover' in df1.columns:
                #df1['Sales Turnover'] = np.round(df1['Sales Turnover'], decimals=2)
                #df1['Sales Turnover'] = (df1['Sales Turnover']*100).astype(int)/100
                df1['Sales Turnover'] = np.round(df1['Sales Turnover'],decimals=6)
                df1['Sales Turnover'] = df1['Sales Turnover'].apply(lambda x: f"{x:.2f}")
            df_list.append(df1)

        except Exception as e:
            print(f"[ERROR] Failed to process {file.filename}: {str(e)}")
            return jsonify({"error": f"Error reading {file.filename}: {str(e)}"}), 500

    # 2. Concatenate ONCE at the end
    if df_list:
        merged_df1 = pd.concat(df_list, ignore_index=True)
    else:
        merged_df1 = pd.DataFrame()

    # 3. Sorting (vectorized)
    primary_sort = request.args.get('primary_sort', 'Distributor')
    secondary_sort = request.args.get('secondary_sort', 'Sales Route')
    primary_asc = request.args.get('primary_asc', 'true').lower() == 'true'
    secondary_asc = request.args.get('secondary_asc', 'true').lower() == 'true'
    sort_cols = [c for c in [primary_sort, secondary_sort] if c in merged_df1.columns]

    if not sort_cols:  # Fallback if sort cols missing
        sorted_df = merged_df1
    else:
        sorted_df = merged_df1.sort_values(
            by=sort_cols,
            ascending=[primary_asc, secondary_asc][:len(sort_cols)]
        )

    # 4. Summary table
    if all(col in sorted_df.columns for col in ['Distributor', 'Distributor Name']):
        summary_df = sorted_df.groupby(['Distributor', 'Distributor Name']).size().reset_index(name='Total Data')
    else:
        summary_df = pd.DataFrame()

    # 5. Fix datetime columns with NaT
    for col in sorted_df.select_dtypes(include=['datetime', 'datetimetz']).columns:
        sorted_df[col] = sorted_df[col].astype(str).replace('NaT', '')

    sorted_df = sorted_df.where(~sorted_df.isna(), '')
    summary_df = summary_df.where(~summary_df.isna(), '')

    print(f"PBI upload processed in {time.time() - t0:.2f} seconds")
    return jsonify({
        "sorted_data_PBI": sorted_df.to_dict(orient='records'),
        "summary_data_PBI": summary_df.to_dict(orient='records')
    })

#-----------------IC EFOS OUTLET---------------------------

@app.route('/upload_ICEFOSOUTLET_OSDP', methods=['POST'])
def upload_files_IC_EFOS_Outlet_OSDP():
    import time
    t0 = time.time()
    if 'files' not in request.files:
        return jsonify({"error": "No files provided"}), 400

    files = request.files.getlist('files')
    if not files or len(files) == 0:
        return jsonify({"error": "Empty file list"}), 400

    # 1. Read all files into DataFrames (in memory)
    df_list = []
    for file in files:
        if file.filename == '':
            continue
        try:
            # Read directly from in-memory stream, not disk
            df = read_excel_mixed(file, usecols=[0,1,2,5,10,11,12,13,14,15,16,21,22])
            df_list.append(df)
        except Exception as e:
            return jsonify({"error": f"Error reading {file.filename}: {str(e)}"}), 500

    # 2. Concatenate ONCE at the end
    if df_list:
        merged_df = pd.concat(df_list, ignore_index=True)
    else:
        merged_df = pd.DataFrame()

    # 3. Data Cleaning
    col_pk_fillin = ['Distributor', 'Distributor Name', 'Sales Route', 'Outlet Code']
    for col in col_pk_fillin:
        if col in merged_df.columns:
            merged_df[col] = merged_df[col].ffill()
    merged_df = merged_df.fillna('')
    if 'Actual Outlet Time(in minutes)' in merged_df.columns:
        merged_df['Actual Outlet Time(in minutes)'] = pd.to_numeric(
            merged_df['Actual Outlet Time(in minutes)'], errors='coerce'
        ).fillna(0)
    if 'Sales Value' in merged_df.columns:
        #merged_df['Sales Value'] = (merged_df['Sales Value'] * 100).astype(int) / 100
        merged_df['Sales Value'] = merged_df['Sales Value'].apply(lambda x: f"{x:.2f}")
    if 'PJP Compliance' in merged_df.columns:
        #merged_df['Sales Value'] = (merged_df['Sales Value'] * 100).astype(int) / 100
        merged_df['PJP Compliance'] = merged_df['PJP Compliance'].replace(r'^\s*$', np.nan, regex=True).fillna(0)

    # 4. Sorting
    primary_sort = request.args.get('primary_sort', 'Distributor')
    secondary_sort = request.args.get('secondary_sort', 'Sales Route')
    primary_asc = request.args.get('primary_asc', 'true').lower() == 'true'
    secondary_asc = request.args.get('secondary_asc', 'true').lower() == 'true'
    sort_cols = [c for c in [primary_sort, secondary_sort] if c in merged_df.columns]
    sorted_df = merged_df.sort_values(by=sort_cols, ascending=[primary_asc, secondary_asc][:len(sort_cols)])

    # 5. Summary
    if all(col in sorted_df.columns for col in ['Distributor', 'Distributor Name']):
        summary_df = sorted_df.groupby(['Distributor', 'Distributor Name']).size().reset_index(name='Total Data')
    else:
        summary_df = pd.DataFrame()

    # 6. Fix datetime columns with NaT
    for col in sorted_df.select_dtypes(include=['datetime', 'datetimetz']).columns:
        sorted_df[col] = sorted_df[col].astype(str).replace('NaT', '')

    sorted_df = sorted_df.rename(
        columns={
        'Time In':'Visit Start Time',     
        'Time Out':'Visit End Time',
        'Actual Outlet Time(in minutes)':'Avg Order Taking Time on Outlet',
        'Effective Outlet Time(in minutes)':'Effective Outlet Time',
        'Sales Value':'Sales Turnover',
        '#SKU':'SKU #'}
    )
    
    sorted_df['Sales Route']= sorted_df['Sales Route'].astype(str).str.strip().str.upper()

    sorted_df = sorted_df.where(~sorted_df.isna(), '')
    print(f"OSDP upload processed in {time.time() - t0:.2f} seconds")

    print("OSDP columns:", sorted_df.columns.tolist())

    return jsonify({
        "sorted_data": sorted_df.to_dict(orient='records'),
        "summary_data": summary_df.to_dict(orient='records')
    })

@app.route('/upload_ICEFOSOUTLET_PBI', methods=['POST'])
def upload_files_IC_EFOS_Outlet_pbi():
    import time
    t0 = time.time()
    if 'files1' not in request.files:
        return jsonify({"error": "No files provided"}), 400

    files = request.files.getlist('files1')
    if not files or len(files) == 0:
        return jsonify({"error": "Empty file list"}), 400

    # 1. Read all files into DataFrames (in memory, no saving to disk)
    df_list = []
    for file in files:
        if file.filename == '':
            continue
        try:
            # Read directly from in-memory stream
            df1 = read_excel_mixed(file, usecols=[0, 1, 2, 5, 10, 11, 12, 13, 14, 15, 16,21,22])
            # Drop last 3 rows (trailing summary rows)
            #if len(df1) >= 3:
            #    df1.drop(df1.tail(3).index, inplace=True)
            df1.columns = [c.strip() for c in df1.columns]
            df1.rename(columns={'# SKU': 'SKU #'}, inplace=True)
            df1 = df1.fillna('')

            # Convert numeric columns if present
            for col in ['SKU #', 'Sales Turnover', 'Avg Order Taking Time on Outlet']:
                if col in df1.columns:
                    df1[col] = pd.to_numeric(df1[col], errors='coerce').fillna(0)

            if 'Distributor' in df1.columns:
                df1['Distributor'] = pd.to_numeric(df1['Distributor'], errors='coerce').fillna(0)
            else:
                return jsonify({"error": f"Missing 'Distributor' column in {file.filename}"}), 400
            
            df1 = df1[df1['Distributor'] != 0]
            # --- Trim by last A/B for this file ---
            if not df1.empty:
                last_A = df1.iloc[-1, 0]
                last_B = df1.iloc[-1, 1]
                if str(last_A).strip() != '' and str(last_B).strip() != '':
                    try:
                        diff = int(float(last_A) - float(last_B))
                        if diff > 0 and diff < len(df1):
                            df1 = df1.iloc[:-diff]
                    except Exception as e:
                        print(f"[WARNING] Couldn't process row-trim by last A/B for {file.filename}: {e}")

            if 'Sales Turnover' in df1.columns:
                #df1['Sales Turnover'] = np.round(df1['Sales Turnover'], decimals=2)
                #df1['Sales Turnover'] = (df1['Sales Turnover']*100).astype(int)/100
                df1['Sales Turnover'] = np.round(df1['Sales Turnover'],decimals=6)
                df1['Sales Turnover'] = df1['Sales Turnover'].apply(lambda x: f"{x:.2f}")

            if 'PJP Compliance' in df1.columns:
                df1['PJP Compliance'] = df1['PJP Compliance'].replace(r'^\s*$', np.nan, regex=True).fillna(0)

            if 'Geo Code Matched' in df1.columns:
                df1['Geo Code Matched'] = df1['Geo Code Matched'].replace(r'^\s*$', np.nan, regex=True).fillna(0)

            df_list.append(df1)

        except Exception as e:
            print(f"[ERROR] Failed to process {file.filename}: {str(e)}")
            return jsonify({"error": f"Error reading {file.filename}: {str(e)}"}), 500

    # 2. Concatenate ONCE at the end
    if df_list:
        merged_df1 = pd.concat(df_list, ignore_index=True)
    else:
        merged_df1 = pd.DataFrame()

    # 3. Sorting (vectorized)
    primary_sort = request.args.get('primary_sort', 'Distributor')
    secondary_sort = request.args.get('secondary_sort', 'Sales Route')
    primary_asc = request.args.get('primary_asc', 'true').lower() == 'true'
    secondary_asc = request.args.get('secondary_asc', 'true').lower() == 'true'
    sort_cols = [c for c in [primary_sort, secondary_sort] if c in merged_df1.columns]

    if not sort_cols:  # Fallback if sort cols missing
        sorted_df = merged_df1
    else:
        sorted_df = merged_df1.sort_values(
            by=sort_cols,
            ascending=[primary_asc, secondary_asc][:len(sort_cols)]
        )

    # 4. Summary table
    if all(col in sorted_df.columns for col in ['Distributor', 'Distributor Name']):
        summary_df = sorted_df.groupby(['Distributor', 'Distributor Name']).size().reset_index(name='Total Data')
    else:
        summary_df = pd.DataFrame()

    # 5. Fix datetime columns with NaT
    for col in sorted_df.select_dtypes(include=['datetime', 'datetimetz']).columns:
        sorted_df[col] = sorted_df[col].astype(str).replace('NaT', '')

    sorted_df = sorted_df.where(~sorted_df.isna(), '')
    summary_df = summary_df.where(~summary_df.isna(), '')

    print("PBI columns:", sorted_df.columns.tolist())

    print(f"PBI upload processed in {time.time() - t0:.2f} seconds")
    return jsonify({
        "sorted_data_PBI": sorted_df.to_dict(orient='records'),
        "summary_data_PBI": summary_df.to_dict(orient='records')
    })

#-----------------IC DSS-----------------------------
@app.route('/upload_ICDSS_OSDP', methods=['POST'])
def upload_files_IC_DSS_OSDP():
    import time
    t0 = time.time()
    if 'files' not in request.files:
        return jsonify({"error": "No files provided"}), 400

    files = request.files.getlist('files')
    if not files or len(files) == 0:
        return jsonify({"error": "Empty file list"}), 400
    
    required_columns = [
        'Distributor',
        'Distributor Name',
        'Sales Route',
        'Outlet Code',
        'Invoice Date',
        'Invoice No',
        'Sale Qty CS',
        'Sale Qty PC',
        'Free Total Qty',
        'GSV(Amount)',
        'NIV(Net Invoice Value)',
        'Sales Turn Over',
    ]

    # 1. Read all files into DataFrames (in memory)
    df_list = []
    for file in files:
        if file.filename == '':
            continue
        try:
            # Read directly from in-memory stream, not disk
            excel_path = file.stream
            sheet_name = 'NGDMS CR Daily Sales Summary'  # Update if different
            df = pd.read_excel(excel_path, sheet_name=sheet_name, usecols=required_columns)
            cols_to_check = ['Distributor', 'Sales Route', 'Outlet Code', 'Invoice Date', 'Invoice No']  # <-- change if different!
            mask = ~df[cols_to_check].apply(lambda row: row.astype(str).str.contains('Total', case=False, na=False)).any(axis=1)
            filtered_df = df[mask].reset_index(drop=True)
            #filtered_df = filtered_df.drop(columns=['Invoice Date'])
            df_list.append(filtered_df)
        except Exception as e:
            import traceback
            print("ERROR while reading file:", file.filename)
            traceback.print_exc()  # <-- This will print the full traceback
            return jsonify({"error": f"Error reading {file.filename}: {str(e)}"}), 500


    # 2. Concatenate ONCE at the end
    if df_list:
        merged_df = pd.concat(df_list, ignore_index=True)
    else:
        merged_df = pd.DataFrame()

    # 3. Data Cleaning
    col_pk_fillin = ['Distributor', 'Distributor Name', 'Sales Route', 'Outlet Code','Invoice Date',]
    for col in col_pk_fillin:
        if col in merged_df.columns:
            merged_df[col] = merged_df[col].ffill()
    merged_df = merged_df.fillna('')

    # 4. Sorting
    primary_sort = request.args.get('primary_sort', 'Distributor')
    secondary_sort = request.args.get('secondary_sort', 'Sales Route')
    primary_asc = request.args.get('primary_asc', 'true').lower() == 'true'
    secondary_asc = request.args.get('secondary_asc', 'true').lower() == 'true'
    sort_cols = [c for c in [primary_sort, secondary_sort] if c in merged_df.columns]
    sorted_df = merged_df.sort_values(by=sort_cols, ascending=[primary_asc, secondary_asc][:len(sort_cols)])

    # 5. Summary
    if all(col in sorted_df.columns for col in ['Distributor', 'Distributor Name']):
        summary_df = sorted_df.groupby(['Distributor', 'Distributor Name']).size().reset_index(name='Total Data')
    else:
        summary_df = pd.DataFrame()

    # 6. Fix datetime columns with NaT, and reformat 'Invoice Date'
    if 'Invoice Date' in sorted_df.columns:
        # Convert to datetime (errors='coerce' turns invalid into NaT)
        sorted_df['Invoice Date'] = pd.to_datetime(sorted_df['Invoice Date'], errors='coerce')
        # Format to YYYY-MM-DD, missing values become empty string
        sorted_df['Invoice Date'] = sorted_df['Invoice Date'].dt.strftime('%Y-%m-%d').fillna('')

    #7. Rename column name to match
    sorted_df['Invoice No']= sorted_df['Invoice No'].astype(str).str.strip().str.upper()
    sorted_df['Sales Route']= sorted_df['Sales Route'].astype(str).str.strip().str.upper()
    sorted_df.rename(columns={'Sales Turn Over': 'Sales Turnover'}, inplace=True)
    sorted_df = sorted_df.where(~sorted_df.isna(), '')

    if 'Outlet Code' in sorted_df.columns:
        sorted_df = sorted_df[~sorted_df['Outlet Code'].astype(str).str.startswith('TOT')]

    print(f"OSDP upload processed in {time.time() - t0:.2f} seconds")
    #print("OSDP DataFrame:\n", sorted_df[['Invoice No','Invoice Date']])  # <--- Print to terminal/console
    return jsonify({
        "sorted_data": sorted_df.to_dict(orient='records'),
        "summary_data": summary_df.to_dict(orient='records')
    })

@app.route('/upload_ICDSS_PBI', methods=['POST'])
def upload_files_IC_DSS_Outlet_pbi():
    import time
    t0 = time.time()
    if 'files1' not in request.files:
        return jsonify({"error": "No files provided"}), 400

    files = request.files.getlist('files1')
    if not files or len(files) == 0:
        return jsonify({"error": "Empty file list"}), 400

    # 1. Read all files into DataFrames (in memory, no saving to disk)
    df_list = []
    for file in files:
        if file.filename == '':
            continue
        try:
            # Read directly from in-memory stream
            df1 = pd.read_excel(file.stream, usecols=[0,1,2,5,8,9,13,14,15,17,23,25], engine='openpyxl')
            # Drop last 3 rows (trailing summary rows)
            #if len(df1) >= 3:
            #    df1.drop(df1.tail(3).index, inplace=True)
            df1.columns = [c.strip() for c in df1.columns]
            df1['Document Number']= df1['Document Number'].astype(str).str.strip().str.upper()
            df1.rename(columns={'Document Number': 'Invoice No'}, inplace=True)
            df1.rename(columns={'Date': 'Invoice Date'}, inplace=True)
            df1 = df1.fillna('')

            if 'Distributor' in df1.columns:
                df1['Distributor'] = pd.to_numeric(df1['Distributor'], errors='coerce').fillna(0)
            else:
                return jsonify({"error": f"Missing 'Distributor' column in {file.filename}"}), 400

            int_columns = ['Sale Qty CS', 'Sale Qty PC', 'Free Total Qty']
            for col in int_columns:
                if col in df1.columns:
                    df1[col] = np.round(pd.to_numeric(df1[col], errors='coerce').fillna(0), 2)  # round to 2 decimals first
                    df1[col] = np.round(df1[col]).astype(int)  # then round to nearest int and cast

            float_columns = ['GSV(Amount)', 'NIV(Net Invoice Value)', 'Sales Turnover']
            for col in float_columns:
                if col in df1.columns:
                    df1[col] = pd.to_numeric(df1[col], errors='coerce').fillna(0.0).astype(float)
            
            df1 = df1[df1['Distributor'] != 0]
            # --- Trim by last A/B for this file ---
            if not df1.empty:
                last_A = df1.iloc[-1, 0]
                last_B = df1.iloc[-1, 1]
                if str(last_A).strip() != '' and str(last_B).strip() != '':
                    try:
                        diff = int(float(last_A) - float(last_B))
                        if diff > 0 and diff < len(df1):
                            df1 = df1.iloc[:-diff]
                    except Exception as e:
                        print(f"[WARNING] Couldn't process row-trim by last A/B for {file.filename}: {e}")

            if 'Invoice No' in df1.columns:
                df1['Invoice No'] = df1['Invoice No'].astype(str).str.strip().str.upper()

            if 'Sales Turnover' in df1.columns:
                df1['Sales Turnover'] = np.round(df1['Sales Turnover'],decimals=4)

            if 'Outlet Code' in df1.columns:
                df1 = df1[~df1['Outlet Code'].astype(str).str.startswith('TOT')]

            df_list.append(df1)

        except Exception as e:
            print(f"[ERROR] Failed to process {file.filename}: {str(e)}")
            return jsonify({"error": f"Error reading {file.filename}: {str(e)}"}), 500

    # 2. Concatenate ONCE at the end
    if df_list:
        merged_df1 = pd.concat(df_list, ignore_index=True)
    else:
        merged_df1 = pd.DataFrame()

    # 3. Sorting (vectorized)
    primary_sort = request.args.get('primary_sort', 'Distributor')
    secondary_sort = request.args.get('secondary_sort', 'Sales Route')
    primary_asc = request.args.get('primary_asc', 'true').lower() == 'true'
    secondary_asc = request.args.get('secondary_asc', 'true').lower() == 'true'
    sort_cols = [c for c in [primary_sort, secondary_sort] if c in merged_df1.columns]

    if not sort_cols:  # Fallback if sort cols missing
        sorted_df = merged_df1
    else:
        sorted_df = merged_df1.sort_values(
            by=sort_cols,
            ascending=[primary_asc, secondary_asc][:len(sort_cols)]
        )

    # 4. Fix datetime columns with NaT
    for col in sorted_df.select_dtypes(include=['datetime', 'datetimetz']).columns:
        sorted_df[col] = sorted_df[col].astype(str).replace('NaT', '')

    # 5. Group by first 5 columns, sum the rest
    key_columns = ['Distributor', 'Distributor Name', 'Sales Route', 'Outlet Code', 'Invoice Date', 'Invoice No']
    value_columns = [col for col in sorted_df.columns if col not in key_columns]


    # Make sure value_columns are numeric, or else .sum() won't work as expected!
    df_grouped = sorted_df.groupby(list(key_columns), as_index=False)[list(value_columns)].sum(numeric_only=True)

    # 6. Summary table
    if all(col in df_grouped.columns for col in ['Distributor', 'Distributor Name']):
        summary_df = df_grouped.groupby(['Distributor', 'Distributor Name']).size().reset_index(name='Total Data')
    else:
        summary_df = pd.DataFrame()


    df_grouped = df_grouped.where(~df_grouped.isna(), '')
    summary_df = summary_df.where(~summary_df.isna(), '')
    
    data = df_grouped.to_dict(orient='records')
    for row in data:
        for col in ['Sales Turnover', 'NIV(Net Invoice Value)', 'GSV(Amount)']:
            if col in row and isinstance(row[col], float):
                row[col] = round(row[col], 4)

    #print("PBI DataFrame:\n", data)  # <--- Print to terminal/console
    print(f"PBI upload processed in {time.time() - t0:.2f} seconds")
    return jsonify({
        "sorted_data_PBI": data,
        "summary_data_PBI": summary_df.to_dict(orient='records')
    })

#---------------IC IQ PERFORMANCE---------------------------

@app.route('/upload_ICIQ_OSDP', methods=['POST'])
def upload_files_IC_IQ_OSDP():
    import time
    import pandas as pd
    t0 = time.time()
    
    if 'files' not in request.files:
        return jsonify({"error": "No files provided"}), 400

    files = request.files.getlist('files')
    if not files or len(files) == 0:
        return jsonify({"error": "Empty file list"}), 400
    
    required_columns = [
        'Distributor',
        'Distributor Name',
        'Sales Route',
        'Outlet Code',
        'Outlet Status',
        'Outlet Sales',
        'Route Sales',
        'COTC Total',
        'COTC Ach',
        'COTC Ach %',
        'EB Total',
        'EB Ach',
        'EB Ach %',
        'RL Total',
        'RL Ach',
        'RL Ach %',
    ]

    df_list = []
    skipped_files = []
    structure_mismatch_files = []

    for file in files:
        if file.filename == '':
            continue
        try:
            excel_path = file.stream
            sheet_name = 'UID IC IQ Perfomance Report'
            # Read just the header row (skiprows=4 as in original)
            test_df = pd.read_excel(excel_path, sheet_name=sheet_name, skiprows=4, nrows=0)
            available_cols = [col.strip() for col in test_df.columns]
            used_cols = [col for col in required_columns if col in available_cols]
            missing_cols = [col for col in required_columns if col not in available_cols]
            
            if len(used_cols) < 5:
                print(f"[SKIPPED: STRUCTURE MISMATCH] File: {file.filename}, found only columns: {used_cols}")
                structure_mismatch_files.append(file.filename)
                continue

            # Seek back to start of stream to read full data
            file.stream.seek(0)
            df = pd.read_excel(file.stream, sheet_name=sheet_name, usecols=used_cols, skiprows=4)

            # If no data, skip file
            if df.empty:
                print(f"[SKIPPED: NO DATA] No data in file: {file.filename}")
                skipped_files.append(file.filename)
                continue

            # Add missing columns as empty if needed
            for col in missing_cols:
                df[col] = ''
            # Reorder columns to required_columns order
            df = df[[col for col in required_columns]]

            df_list.append(df)
        except Exception as e:
            import traceback
            print("ERROR while reading file:", file.filename)
            traceback.print_exc()
            return jsonify({"error": f"Error reading {file.filename}: {str(e)}"}), 500

    if df_list:
        merged_df = pd.concat(df_list, ignore_index=True)
    else:
        merged_df = pd.DataFrame(columns=required_columns)

    col_pk_fillin = ['Distributor', 'Distributor Name', 'Sales Route']
    for col in col_pk_fillin:
        if col in merged_df.columns:
            merged_df[col] = merged_df[col].ffill()

    #if 'Outlet Status' in merged_df.columns:
        #merged_df = merged_df[~merged_df['Outlet Status'].astype(str).str.startswith('In')]
        #merged_df = merged_df.drop(columns=['Outlet Status'])
    merged_df = merged_df.fillna('')

    primary_sort = request.args.get('primary_sort', 'Distributor')
    secondary_sort = request.args.get('secondary_sort', 'Sales Route')
    primary_asc = request.args.get('primary_asc', 'true').lower() == 'true'
    secondary_asc = request.args.get('secondary_asc', 'true').lower() == 'true'
    sort_cols = [c for c in [primary_sort, secondary_sort] if c in merged_df.columns]
    sorted_df = merged_df.sort_values(by=sort_cols, ascending=[primary_asc, secondary_asc][:len(sort_cols)])

    if all(col in sorted_df.columns for col in ['Distributor', 'Distributor Name']):
        summary_df = sorted_df.groupby(['Distributor', 'Distributor Name']).size().reset_index(name='Total Data')
    else:
        summary_df = pd.DataFrame()

    if 'Invoice Date' in sorted_df.columns:
        sorted_df['Invoice Date'] = pd.to_datetime(sorted_df['Invoice Date'], errors='coerce')
        sorted_df['Invoice Date'] = sorted_df['Invoice Date'].dt.strftime('%Y-%m-%d').fillna('')

    if 'Outlet Code' in sorted_df.columns:
    # Exclude null (NaN) and empty Outlet Code, and those starting with 'TOT'
        sorted_df = sorted_df[
            sorted_df['Outlet Code'].notna() &                               # Not NaN
            (sorted_df['Outlet Code'].astype(str).str.strip() != '') &       # Not empty string
            ~sorted_df['Outlet Code'].astype(str).str.startswith('TOT')      # Not starts with 'TOT'
        ]

    if 'Sales Route' in sorted_df.columns:
        sorted_df['Sales Route'] = sorted_df['Sales Route'].astype(str).str.strip().str.upper()

    print(f"OSDP upload processed in {time.time() - t0:.2f} seconds")
    if skipped_files:
        print("[SKIPPED: NO DATA FILES]", skipped_files)
    if structure_mismatch_files:
        print("[SKIPPED: STRUCTURE MISMATCH FILES]", structure_mismatch_files)

    return jsonify({
        "sorted_data": sorted_df.to_dict(orient='records'),
        "summary_data": summary_df.to_dict(orient='records'),
        "skipped_files": skipped_files,
        "structure_mismatch_files": structure_mismatch_files,
    })


@app.route('/upload_ICIQ_PBI', methods=['POST'])
def upload_files_IC_IQ_Outlet_pbi():
    import time
    t0 = time.time()
    if 'files1' not in request.files:
        return jsonify({"error": "No files provided"}), 400

    files = request.files.getlist('files1')
    if not files or len(files) == 0:
        return jsonify({"error": "Empty file list"}), 400
    
    required_columns = [
        'Distributor',
        'Distributor Name',
        'Sales Route',
        'Outlet Code',
        'Outlet Active',
        'Outlet Sales',
        'Route Sales',
        'COTC Total',
        'COTC Ach',
        '% COTC Ach',
        'EB Total',
        'EB Ach',
        '% EB Ach',
        'RL Total',
        'RL Ach',
        '% RL Ach',
    ]

    # 1. Read all files into DataFrames (in memory, no saving to disk)
    df_list = []
    for file in files:
        if file.filename == '':
            continue
        try:
            # Read directly from in-memory stream
            df1 = pd.read_excel(file.stream, usecols=required_columns, engine='openpyxl')
            df1.columns = [c.strip() for c in df1.columns]
            df1.rename(columns={'% COTC Ach': 'COTC Ach %'}, inplace=True)
            df1.rename(columns={'Date': 'Invoice Date'}, inplace=True)
            df1 = df1.fillna('')

            if 'Distributor' in df1.columns:
                df1['Distributor'] = pd.to_numeric(df1['Distributor'], errors='coerce').fillna(0)
            else:
                return jsonify({"error": f"Missing 'Distributor' column in {file.filename}"}), 400
            
            df1 = df1[df1['Distributor'] != 0]
            # --- Trim by last A/B for this file ---
            if not df1.empty:
                last_A = df1.iloc[-1, 0]
                last_B = df1.iloc[-1, 1]
                if str(last_A).strip() != '' and str(last_B).strip() != '':
                    try:
                        diff = int(float(last_A) - float(last_B))
                        if diff > 0 and diff < len(df1):
                            df1 = df1.iloc[:-diff]
                    except Exception as e:
                        print(f"[WARNING] Couldn't process row-trim by last A/B for {file.filename}: {e}")


            if 'Outlet Code' in df1.columns:
                df1 = df1[~df1['Outlet Code'].astype(str).str.startswith('TOT')]


            #if 'Outlet Active' in df1.columns:
                #df1 = df1[~df1['Outlet Active'].astype(str).str.startswith('No')]

            #df1 = df1.drop(columns=['Outlet Active'])
            df_list.append(df1)

        except Exception as e:
            print(f"[ERROR] Failed to process {file.filename}: {str(e)}")
            return jsonify({"error": f"Error reading {file.filename}: {str(e)}"}), 500

    # 2. Concatenate ONCE at the end
    if df_list:
        merged_df1 = pd.concat(df_list, ignore_index=True)
    else:
        merged_df1 = pd.DataFrame()

    # 3. Sorting (vectorized)
    primary_sort = request.args.get('primary_sort', 'Distributor')
    secondary_sort = request.args.get('secondary_sort', 'Sales Route')
    primary_asc = request.args.get('primary_asc', 'true').lower() == 'true'
    secondary_asc = request.args.get('secondary_asc', 'true').lower() == 'true'
    sort_cols = [c for c in [primary_sort, secondary_sort] if c in merged_df1.columns]

    if not sort_cols:  # Fallback if sort cols missing
        sorted_df = merged_df1
    else:
        sorted_df = merged_df1.sort_values(
            by=sort_cols,
            ascending=[primary_asc, secondary_asc][:len(sort_cols)]
        )

    # 4. Fix datetime columns with NaT
    for col in sorted_df.select_dtypes(include=['datetime', 'datetimetz']).columns:
        sorted_df[col] = sorted_df[col].astype(str).replace('NaT', '')

    # 6. Summary table
    if all(col in sorted_df.columns for col in ['Distributor', 'Distributor Name']):
        summary_df = sorted_df.groupby(['Distributor', 'Distributor Name']).size().reset_index(name='Total Data')
    else:
        summary_df = pd.DataFrame()


    sorted_df = sorted_df.where(~sorted_df.isna(), '')
    summary_df = summary_df.where(~summary_df.isna(), '')


    #print("PBI DataFrame:\n", data)  # <--- Print to terminal/console
    print(f"PBI upload processed in {time.time() - t0:.2f} seconds")
    return jsonify({
        "sorted_data_PBI": sorted_df.to_dict(orient='records'),
        "summary_data_PBI": summary_df.to_dict(orient='records')
    })

#-----------------------------Raw Data Invoive Level------------------------------------------
@app.route('/upload_HPCRAWDATA_OSDP', methods=['POST'])
def upload_files_hpc_rawdata_OSDP():
    import time
    import pandas as pd
    t0 = time.time()

    if 'files' not in request.files:
        return jsonify({"error": "No files provided"}), 400

    files = request.files.getlist('files')
    if not files or len(files) == 0:
        return jsonify({"error": "Empty file list"}), 400

    required_columns = [
        'Distributor Code', 'Distributor Name', 'Pop Type', 'Outlet Code', 'Outlet Name',
        'INVH_NO', 'INVH_DATE', #'Lead Based Pack',
        'Prod Code', 'Prod Name',
        'GSV', 'Net Amount', 'Billing Type', 'INVTYPE', 'Cases', 'List Price Per Case',
    ]
    required_columns1 = [
        'Distributor Code', 'Outlet Code', 'INVH_NO', 'Prod Code', 'INVTYPE', 'Cases',
    ]

    df_list = []
    skipped_files = []
    structure_mismatch_files = []
    debug_columns_info = {}

    for file in files:
        if file.filename == '':
            continue
        try:
            # Always start at beginning of stream
            file.stream.seek(0)
            # Read with actual header row (row 0)
            df = pd.read_excel(file.stream, sheet_name=0, header=1)
            df.columns = [str(col).strip() for col in df.columns]

            available_cols = list(df.columns)
            used_cols = [col for col in required_columns if col in available_cols]
            missing_cols = [col for col in required_columns if col not in available_cols]
            debug_columns_info[file.filename] = {"available_cols": available_cols, "missing_cols": missing_cols}

            if len(used_cols) < 5:
                print(f"[SKIPPED: STRUCTURE MISMATCH] File: {file.filename}, found only columns: {used_cols}")
                structure_mismatch_files.append(file.filename)
                continue

            if df.empty:
                print(f"[SKIPPED: NO DATA] No data in file: {file.filename}")
                skipped_files.append(file.filename)
                continue

            # Add missing columns as empty if needed
            for col in missing_cols:
                df[col] = ''

            df['INVH_NO'] = df['INVH_NO'].astype(str).str[:12]
            df['Prod Code'] = df['Prod Code'].astype(str)
            # Select required columns and create copies
            df1 = df[required_columns].copy()
            df2 = df[required_columns1].copy()
            # Filter out 'FREE'
            df1 = df1[~df1['INVTYPE'].astype(str).str.upper().str.startswith('FREE')]
            df2 = df2[df2['INVTYPE'].astype(str).str.upper().str.startswith('FREE')]
            df1 = df1.drop(columns='INVTYPE')
            df2 = df2.drop(columns='INVTYPE')

            merged_df1 = pd.merge(
                df1, df2, on=['Distributor Code', 'Outlet Code', 'INVH_NO', 'Prod Code'], how='left'
            )
            merged_df1['Cases_y'] = merged_df1['Cases_y'].fillna(0)
            merged_df1 = merged_df1.rename(
                columns={'Cases_x': 'Cases', 'Cases_y': 'Free QTY in CS', 'Distributor Code': 'Distributor','INVH_NO':'Invoice No','INVH_DATE':'Invoice Date'}
            )
            df_list.append(merged_df1)
            
        except Exception as e:
            import traceback
            print(f"ERROR while reading file: {file.filename}")
            traceback.print_exc()
            debug_columns_info[file.filename] = {"error": str(e)}
            return jsonify({
                "error": f"Error reading {file.filename}: {str(e)}",
                "debug_columns_info": debug_columns_info
            }), 500

    if df_list:
        merged_df = pd.concat(df_list, ignore_index=True)
    else:
        merged_df = pd.DataFrame(columns=required_columns)

    merged_df = merged_df.fillna('')
    # Sorting and summary
    primary_sort = request.args.get('primary_sort', 'Distributor')
    secondary_sort = request.args.get('secondary_sort', 'Sales Route')
    primary_asc = request.args.get('primary_asc', 'true').lower() == 'true'
    secondary_asc = request.args.get('secondary_asc', 'true').lower() == 'true'
    sort_cols = [c for c in [primary_sort, secondary_sort] if c in merged_df.columns]
    sorted_df = merged_df.sort_values(by=sort_cols, ascending=[primary_asc, secondary_asc][:len(sort_cols)]) if not merged_df.empty else merged_df

    if all(col in sorted_df.columns for col in ['Distributor', 'Distributor Name']):
        summary_df = sorted_df.groupby(['Distributor', 'Distributor Name']).size().reset_index(name='Total Data')
    else:
        summary_df = pd.DataFrame()

    if 'INVH_DATE' in sorted_df.columns:
        sorted_df['INVH_DATE'] = pd.to_datetime(sorted_df['INVH_DATE'], errors='coerce')
        sorted_df['INVH_DATE'] = sorted_df['INVH_DATE'].dt.strftime('%Y-%m-%d').fillna('')

    if 'Sales Route' in sorted_df.columns:
        sorted_df['Sales Route'] = sorted_df['Sales Route'].astype(str).str.strip().str.upper()
    
    sorted_df['Free QTY in CS'] = sorted_df['Free QTY in CS'].apply(lambda x: f"{x:.8f}")
    sorted_df['Cases'] = sorted_df['Cases'].apply(lambda x: f"{x:.8f}")

    print(f"OSDP upload processed in {time.time() - t0:.2f} seconds")
    if skipped_files:
        print("[SKIPPED: NO DATA FILES]", skipped_files)
    if structure_mismatch_files:
        print("[SKIPPED: STRUCTURE MISMATCH FILES]", structure_mismatch_files)
    print("merged_df shape:", merged_df.shape)
    print("merged_df columns:", list(merged_df.columns))
    print("merged_df head:", merged_df.head())

    # Extra debug info for frontend
    if merged_df.empty:
        return jsonify({
            "sorted_data": [],
            "summary_data": [],
            "skipped_files": skipped_files,
            "structure_mismatch_files": structure_mismatch_files,
            "debug_columns_info": debug_columns_info,
            "error": "Merged DataFrame is empty. See debug_columns_info for column issues."
        }), 200

    return jsonify({
        "sorted_data": sorted_df.to_dict(orient='records'),
        "summary_data": summary_df.to_dict(orient='records'),
        "debug_columns_info": debug_columns_info
    })

@app.route('/upload_HPCRAWDATA_PBI', methods=['POST'])
def upload_files_HPC_RAW_data_pbi():
    import time
    t0 = time.time()
    if 'files1' not in request.files:
        return jsonify({"error": "No files provided"}), 400

    files = request.files.getlist('files1')
    if not files or len(files) == 0:
        return jsonify({"error": "Empty file list"}), 400
    
    required_columns = [
        'Distributor',
        'Distributor Name',
        'Outlet Sub Type (CO 5) ',
        'Outlet Code',
        'Outlet Name',
        'Invoice No',
        'Invoice Date',
        #'Base Pack Description',
        'Product ID',
        'Product Description',
        'GSV(Amount) ',
        'NIV (Net Invoice Value )',
        'Billing Type',
        'Pack Size',
        'Total Qty',
        'Free Total Qty',
        'Price',
    ]

    # 1. Read all files into DataFrames (in memory, no saving to disk)
    df_list = []
    for file in files:
        if file.filename == '':
            continue
        try:
            # Read directly from in-memory stream
            df1 = pd.read_excel(file.stream, usecols=required_columns, engine='openpyxl')
            df1.columns = [c.strip() for c in df1.columns]
            df1['Cases']=np.trunc((df1['Total Qty']/df1['Pack Size'])*1e8)/1e8
            df1['List Price Per Case']=df1['Price']*df1['Pack Size']
            df1['Free QTY in CS'] = np.trunc((df1['Free Total Qty'] / df1['Pack Size']) * 1e8) / 1e8
            df1= df1[
                df1['Price'].notna() &                               # Not NaN
                (df1['Price'].astype(str).str.strip() != '')         # Not empty string
            ]
            print(df1.columns)
            df1 = df1.drop(columns=['Total Qty', 'Pack Size', 'Price', 'Free Total Qty'])
            df1 = df1.rename(
                columns={'Product ID': 'Prod Code', 'Product Description': 'Prod Name', 'GSV(Amount)': 'GSV','NIV (Net Invoice Value )':'Net Amount','Outlet Sub Type (CO 5)':'POP Type',
                        'Base Pack Description': 'Lead Based Pack'}
            )
            df1 = df1.fillna('')

            if 'Distributor' in df1.columns:
                df1['Distributor'] = pd.to_numeric(df1['Distributor'], errors='coerce').fillna(0)
            else:
                return jsonify({"error": f"Missing 'Distributor' column in {file.filename}"}), 400
            
            df1 = df1[df1['Distributor'] != 0]
            # --- Trim by last A/B for this file ---
            if not df1.empty:
                last_A = df1.iloc[-1, 0]
                last_B = df1.iloc[-1, 1]
                if str(last_A).strip() != '' and str(last_B).strip() != '':
                    try:
                        diff = int(float(last_A) - float(last_B))
                        if diff > 0 and diff < len(df1):
                            df1 = df1.iloc[:-diff]
                    except Exception as e:
                        print(f"[WARNING] Couldn't process row-trim by last A/B for {file.filename}: {e}")


            if 'Outlet Code' in df1.columns:
                df1 = df1[~df1['Outlet Code'].astype(str).str.startswith('TOT')]


            #if 'Outlet Active' in df1.columns:
                #df1 = df1[~df1['Outlet Active'].astype(str).str.startswith('No')]

            #df1 = df1.drop(columns=['Outlet Active'])
            df_list.append(df1)

        except Exception as e:
            print(f"[ERROR] Failed to process {file.filename}: {str(e)}")
            return jsonify({"error": f"Error reading {file.filename}: {str(e)}"}), 500

    # 2. Concatenate ONCE at the end
    if df_list:
        merged_df1 = pd.concat(df_list, ignore_index=True)
    else:
        merged_df1 = pd.DataFrame()

    merged_df1['List Price Per Case']=merged_df1['List Price Per Case'].round(2)
    merged_df1['Free QTY in CS'] = merged_df1['Free QTY in CS'].apply(lambda x: f"{x:.8f}")
    merged_df1['Cases'] = merged_df1['Cases'].apply(lambda x: f"{x:.8f}")


    # 3. Sorting (vectorized)
    primary_sort = request.args.get('primary_sort', 'Distributor')
    secondary_sort = request.args.get('secondary_sort', 'Sales Route')
    primary_asc = request.args.get('primary_asc', 'true').lower() == 'true'
    secondary_asc = request.args.get('secondary_asc', 'true').lower() == 'true'
    sort_cols = [c for c in [primary_sort, secondary_sort] if c in merged_df1.columns]

    if not sort_cols:  # Fallback if sort cols missing
        sorted_df = merged_df1
    else:
        sorted_df = merged_df1.sort_values(
            by=sort_cols,
            ascending=[primary_asc, secondary_asc][:len(sort_cols)]
        )

    # 4. Fix datetime columns with NaT
    for col in sorted_df.select_dtypes(include=['datetime', 'datetimetz']).columns:
        sorted_df[col] = sorted_df[col].astype(str).replace('NaT', '')

    # 6. Summary table
    if all(col in sorted_df.columns for col in ['Distributor', 'Distributor Name']):
        summary_df = sorted_df.groupby(['Distributor', 'Distributor Name']).size().reset_index(name='Total Data')
    else:
        summary_df = pd.DataFrame()


    sorted_df = sorted_df.where(~sorted_df.isna(), '')
    summary_df = summary_df.where(~summary_df.isna(), '')

    print(sorted_df.columns)
    #print("PBI DataFrame:\n", data)  # <--- Print to terminal/console
    print(f"PBI upload processed in {time.time() - t0:.2f} seconds")
    return jsonify({
        "sorted_data_PBI": sorted_df.to_dict(orient='records'),
        "summary_data_PBI": summary_df.to_dict(orient='records')
    })


##---------------------------Mismatch Tracker Report----------------------------------------------

@app.route('/api/mismatches', methods=['GET'])
def get_mismatches():
    url = "https://docs.google.com/spreadsheets/d/1ql1BfkiuRuU3A3mfOxEw_GoL2gP5ki7eQECHxyfvFwk/export?format=csv&gid=0"
    df = pd.read_csv(url)
    mismatches = df[df['Report Status'].astype(str).str.lower().str.strip() == 'mismatch']
    return jsonify(mismatches.to_dict(orient="records"))

##-------------------------DSS Report-----------------------------------------
@app.route("/download-report", methods=["GET"])
def download_report():
    # ---- Step 1: Prepare Raw Data ----
    data = [
        {"DT Code": 15081034, "DT Name": "SWANG CHAI CHUAN SDN BHD", 
         "CSDP Sales Qty CS (Include To)": 20081, "CSDP Sales Qty CS (Exclude To)": 17003,
         "OSDP Sales Qty CS": 17003},
        {"DT Code": 15081169, "DT Name": "SUN CHUAN (LGK) SDN. BHD.", 
         "CSDP Sales Qty CS (Include To)": 5746, "CSDP Sales Qty CS (Exclude To)": 5260,
         "OSDP Sales Qty CS": 5260},
    ]
    df = pd.DataFrame(data)

    # ---- Step 2: Save DataFrame to Excel (without variance yet) ----
    output = BytesIO()
    df.to_excel(output, index=False, startrow=0)
    output.seek(0)

    # ---- Step 3: Insert formulas with openpyxl ----
    wb = load_workbook(output)
    ws = wb.active

    # Add headers
    col_with_to = ws.max_column + 1
    col_without_to = ws.max_column + 2
    ws.cell(row=1, column=col_with_to, value="Variance (With TO)")
    ws.cell(row=1, column=col_without_to, value="Variance (Without TO)")

    # Insert formulas row by row
    for row in range(2, ws.max_row + 1):
        include_col = list(df.columns).index("CSDP Sales Qty CS (Include To)") + 1
        exclude_col = list(df.columns).index("CSDP Sales Qty CS (Exclude To)") + 1
        osdp_col    = list(df.columns).index("OSDP Sales Qty CS") + 1

        ws.cell(row=row, column=col_with_to,
                value=f"={chr(64+include_col)}{row}-{chr(64+osdp_col)}{row}")
        ws.cell(row=row, column=col_without_to,
                value=f"={chr(64+exclude_col)}{row}-{chr(64+osdp_col)}{row}")

    # ---- Step 4: Apply Yellow Fill for Sales Qty ----
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    for col in ["CSDP Sales Qty CS (Include To)", "CSDP Sales Qty CS (Exclude To)", "OSDP Sales Qty CS"]:
        col_idx = list(df.columns).index(col) + 1
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_idx).fill = yellow_fill

    # ---- Step 5: Conditional Formatting for Variances ----
    # Variance (With TO)
    col_letter_with_to = chr(64 + col_with_to)
    range_with_to = f"{col_letter_with_to}2:{col_letter_with_to}{ws.max_row}"

    ws.conditional_formatting.add(
        range_with_to,
        FormulaRule(
            formula=[f"{col_letter_with_to}2=0"],
            stopIfTrue=True,
            font=Font(color="006100"),  # dark green text
            fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # light green
        )
    )
    ws.conditional_formatting.add(
        range_with_to,
        FormulaRule(
            formula=[f"{col_letter_with_to}2<>0"],
            stopIfTrue=True,
            font=Font(color="9C0006"),  # dark red text
            fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # light red
        )
    )

    # Variance (Without TO) — same logic
    col_letter_without_to = chr(64 + col_without_to)
    range_without_to = f"{col_letter_without_to}2:{col_letter_without_to}{ws.max_row}"

    ws.conditional_formatting.add(
        range_without_to,
        FormulaRule(
            formula=[f"{col_letter_without_to}2=0"],
            stopIfTrue=True,
            font=Font(color="006100"),
            fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        )
    )
    ws.conditional_formatting.add(
        range_without_to,
        FormulaRule(
            formula=[f"{col_letter_without_to}2<>0"],
            stopIfTrue=True,
            font=Font(color="9C0006"),
            fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        )
    )
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            # Skip Distributor Code (first column)
            if cell.column == 1:
                continue
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

    # ---- Step 6: Save workbook to memory ----
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="sales_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

#---------------DSS CSV FILE---------------#
@app.route('/export_combined_csv', methods=['POST','GET'])
def export_combined_csv():
    if request.method == 'GET':
        return '''
        <form method="POST" enctype="multipart/form-data">
          <p>OSDP files: <input type="file" name="files" multiple></p>
          <p>PBI files: <input type="file" name="files1" multiple></p>
          <button type="submit">Generate CSV</button>
        </form>
        '''
    """
    Upload OSDP and PBI files together (multipart/form-data):
      - OSDP files  -> field name: 'files'
      - PBI files   -> field name: 'files1'
    Returns: combined CSV (side-by-side) merged on 'Distributor'.
    """
    # --- Validate presence of both sets ---
    if 'files' not in request.files or len(request.files.getlist('files')) == 0:
        return jsonify({"error": "No OSDP files provided (field 'files')"}), 400
    if 'files1' not in request.files or len(request.files.getlist('files1')) == 0:
        return jsonify({"error": "No PBI files provided (field 'files1')"}), 400

    osdp_files = request.files.getlist('files')
    pbi_files  = request.files.getlist('files1')

    # ========== Helpers ==========
    def to_numeric_safe(df, cols):
        for c in cols:
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
        return df

    def build_osdp_export(files):
        """Recreate export_data for OSDP: sum by Distributor."""
        required_columns = [
            'Distributor',
            'Distributor Name',
            'Sales Route',
            'Outlet Code',
            'Invoice Date',
            'Invoice No',
            'Sale Qty CS',
            'Sale Qty PC',
            'Free Total Qty',
            'GSV(Amount)',
            'NIV(Net Invoice Value)',
            'Sales Turn Over',
        ]
        df_list = []
        for f in files:
            if not f or f.filename == '':
                continue
            # Read the named sheet
            df = pd.read_excel(
                f.stream,
                sheet_name='NGDMS CR Daily Sales Summary',
                usecols=required_columns
            )
            # Remove "Total" rows in key cols
            cols_to_check = ['Distributor', 'Sales Route', 'Outlet Code', 'Invoice Date', 'Invoice No']
            mask = ~df[cols_to_check].apply(lambda r: r.astype(str).str.contains('Total', case=False, na=False)).any(axis=1)
            df = df[mask].reset_index(drop=True)
            df_list.append(df)

        if not df_list:
            return pd.DataFrame(columns=['Distributor'])

        df = pd.concat(df_list, ignore_index=True)

        # Basic cleaning & normalization
        for col in ['Distributor', 'Distributor Name', 'Sales Route', 'Outlet Code', 'Invoice Date']:
            if col in df.columns:
                df[col] = df[col].ffill()
        df = df.fillna('')

        # Normalize, rename & drop TOT*
        if 'Invoice No' in df.columns:
            df['Invoice No'] = df['Invoice No'].astype(str).str.strip().str.upper()
        if 'Sales Route' in df.columns:
            df['Sales Route'] = df['Sales Route'].astype(str).str.strip().str.upper()
        df.rename(columns={'Sales Turn Over': 'Sales Turnover'}, inplace=True)
        if 'Outlet Code' in df.columns:
            df = df[~df['Outlet Code'].astype(str).str.startswith('TOT')]

        # Ensure numeric for aggregation
        numeric_targets = [
            'Sale Qty CS', 'Sale Qty PC', 'Free Total Qty',
            'GSV(Amount)', 'NIV(Net Invoice Value)', 'Sales Turnover'
        ]
        df = to_numeric_safe(df, numeric_targets)

        # Group -> sum by Distributor
        export = (
            df.groupby('Distributor', as_index=False)[numeric_targets]
              .sum()
        )

        # Rename back to match your OSDP export labels (keep Sales Turn Over spelling)
        export.rename(columns={'Sales Turnover': 'Sales Turn Over'}, inplace=True)

        return export

    def build_pbi_export(files):
        """Recreate export_data_PBI for PBI: sum by Distributor."""
        usecols = [0,1,2,5,8,9,13,14,15,17,23,25]
        df_list = []
        for f in files:
            if not f or f.filename == '':
                continue
            df1 = pd.read_excel(f.stream, usecols=usecols, engine='openpyxl')
            df1.columns = [c.strip() for c in df1.columns]

            # Normalize column names
            if 'Document Number' in df1.columns:
                df1['Document Number'] = df1['Document Number'].astype(str).str.strip().str.upper()
                df1.rename(columns={'Document Number': 'Invoice No'}, inplace=True)
            if 'Date' in df1.columns:
                df1.rename(columns={'Date': 'Invoice Date'}, inplace=True)

            df1 = df1.fillna('')

            if 'Distributor' not in df1.columns:
                # Skip this file if missing critical key
                continue
            df1['Distributor'] = pd.to_numeric(df1['Distributor'], errors='coerce').fillna(0)
            df1 = df1[df1['Distributor'] != 0]

            # Clean numerics
            for c in ['Sale Qty CS', 'Sale Qty PC', 'Free Total Qty']:
                if c in df1.columns:
                    df1[c] = np.round(pd.to_numeric(df1[c], errors='coerce').fillna(0), 2)
                    df1[c] = np.round(df1[c]).astype(int)

            for c in ['GSV(Amount)', 'NIV(Net Invoice Value)', 'Sales Turnover']:
                if c in df1.columns:
                    df1[c] = pd.to_numeric(df1[c], errors='coerce').fillna(0.0).astype(float)

            if 'Invoice No' in df1.columns:
                df1['Invoice No'] = df1['Invoice No'].astype(str).str.strip().str.upper()
            if 'Sales Turnover' in df1.columns:
                df1['Sales Turnover'] = np.round(df1['Sales Turnover'], 4)
            if 'Outlet Code' in df1.columns:
                df1 = df1[~df1['Outlet Code'].astype(str).str.startswith('TOT')]

            df_list.append(df1)

        if not df_list:
            return pd.DataFrame(columns=['Distributor'])

        df = pd.concat(df_list, ignore_index=True)

        # Targets to sum
        numeric_targets = [
            'Sale Qty CS', 'Sale Qty PC', 'Free Total Qty',
            'GSV(Amount)', 'NIV(Net Invoice Value)', 'Sales Turnover'
        ]
        df = to_numeric_safe(df, numeric_targets)

        export = (
            df.groupby('Distributor', as_index=False)[numeric_targets]
              .sum()
        )
        return export

    # ========== Build both exports ==========
    export_osdp = build_osdp_export(osdp_files)    # columns: Distributor + targets (… Sales Turn Over)
    export_pbi  = build_pbi_export(pbi_files)      # columns: Distributor + targets (… Sales Turnover)

    # Align naming for a clean merge: keep both variants, then suffix
    # First, rename OSDP 'Sales Turn Over' to a neutral form for merging, then suffix later
    osdp_for_merge = export_osdp.copy()
    if 'Sales Turn Over' in osdp_for_merge.columns:
        osdp_for_merge.rename(columns={'Sales Turn Over': 'Sales Turnover'}, inplace=True)

    # Merge on Distributor with suffixes
    merged = pd.merge(
        osdp_for_merge,
        export_pbi,
        on='Distributor',
        how='outer',
        suffixes=('_OSDP', '_PBI')
    )

    # Optional: sort by Distributor
    merged = merged.sort_values(by='Distributor', kind='stable').reset_index(drop=True)

    # Ensure numeric columns are numeric & fill NaNs with 0 for CSV readability
    for c in merged.columns:
        if c != 'Distributor':
            merged[c] = pd.to_numeric(merged[c], errors='coerce').fillna(0)

    # Build CSV in-memory
    buffer = BytesIO()
    merged.to_csv(buffer, index=False).encode('utf-8')
    buffer.seek(0)

    # Send as downloadable CSV
    return send_file(
        buffer,
        mimetype='text/csv',
        as_attachment=True,
        download_name='combined_export.csv'
    )


### PROMOTION AUTO PART########################################################
# simple in-memory cache for current processed data
PROMO_CACHE = {}


def process_monthly_ic_promotion(file_obj):
    """
    Python port of your CommandButton29_Click VBA macro.
    - file_obj: uploaded file from Flask (request.files['file']) or a file path.
    Returns:
        ic_main_df      : processed 'IC Main' equivalent
        ic_sku_df       : processed 'IC SKU' equivalent
        ic_main_out_df  : final U:Z-style summary table
    All kept in-memory (no Excel writing).
    """

    # ---------- Read source workbook (Sheet1) ----------
    src = pd.read_excel(file_obj, sheet_name=0, header=0)

    if src.empty:
        raise ValueError("Source workbook is empty.")

    # Helper: emulate 'last used row' based on a column index
    def last_used_row_by_col(df, col_idx):
        if col_idx >= df.shape[1]:
            return 0
        col = df.iloc[:, col_idx]
        not_null = col[~col.isna()]
        if not not_null.empty:
            # +1 -> convert index to VBA-style row number
            return int(not_null.index[-1]) + 1
        return 0

    # VBA: max(last row in col D, col N)
    last_row_d = last_used_row_by_col(src, 3)   # D
    last_row_n = last_used_row_by_col(src, 13)  # N
    last_row = max(last_row_d, last_row_n)
    if last_row == 0:
        raise ValueError("No data found in D or N columns.")

    # Trim to used rows only
    src = src.iloc[:last_row, :].copy()

    # ==================================================
    # ===============   IC MAIN LOGIC   ================
    # ==================================================
    ic_main = pd.concat([src.iloc[:,:13],src.iloc[:,15:]],axis=1).copy()
    d_col = 3          # col D
    o_col = 14         # col O
    a_col, b_col, c_col = 0, 1, 2

    # 1) Delete rows where D(i) and D(i-1) are both blank, from bottom, starting row 5
    to_drop = []
    for i in range(len(ic_main) - 1, 4 - 1, -1):  # row5 -> index4
        if pd.isna(ic_main.iat[i, d_col]) and pd.isna(ic_main.iat[i - 1, d_col]):
            to_drop.append(ic_main.index[i])
    if to_drop:
        ic_main = ic_main.drop(to_drop)

    ic_main = ic_main.reset_index(drop=True)

    # 2) Numbering promo groups based on column O -> P_seq, Q_group
    col_p = "P_seq"      # like column P in VBA
    col_q = "Q_group"    # like column Q in VBA
    ic_main[col_p] = pd.NA
    ic_main[col_q] = pd.NA

    k = 1  # sequence inside group
    l = 1  # group id
    for idx in range(3, len(ic_main)):  # VBA j = 5 ...##copyni
        val_o = ic_main.iloc[idx, o_col] if o_col < ic_main.shape[1] else pd.NA
        if pd.isna(val_o) or val_o == "":
            k = 1
            l += 1
        else:
            ic_main.at[idx, col_q] = l
            ic_main.at[idx, col_p] = k
            k += 1

    # 3) Delete rows where D is blank
    ic_main = ic_main[~ic_main.iloc[:, d_col].isna()].reset_index(drop=True)

    # 4) Fill down A,B,C when B is blank but D not blank
    for idx in range(1, len(ic_main)):
        if pd.isna(ic_main.iat[idx, b_col]) and not pd.isna(ic_main.iat[idx, d_col]):
            ic_main.iat[idx, a_col] = ic_main.iat[idx - 1, a_col]
            ic_main.iat[idx, b_col] = ic_main.iat[idx - 1, b_col]
            ic_main.iat[idx, c_col] = ic_main.iat[idx - 1, c_col]

    # ==================================================
    # ===============    IC SKU LOGIC   ================
    # ==================================================
    # Here we mirror your pattern: clean & group rows.
    ic_sku = src.iloc[:, [13, 14]].copy() ##copyni
    a_col_sku = 0

    # 1) Delete rows where A(i) and A(i-1) blank from bottom, starting row 5
    to_drop = []
    for i in range(len(ic_sku) - 1, 4 - 1, -1):
        if pd.isna(ic_sku.iat[i, a_col_sku]) and pd.isna(ic_sku.iat[i - 1, a_col_sku]):
            to_drop.append(ic_sku.index[i])
    if to_drop:
        ic_sku = ic_sku.drop(to_drop)

    ic_sku = ic_sku.reset_index(drop=True)

    # 2) Number into group/sequence columns (equivalent C & D)
    col_c = "SKU_seq_in_group"
    col_d = "SKU_group"
    ic_sku[col_c] = pd.NA
    ic_sku[col_d] = pd.NA

    c = 1
    d = 1
    for idx in range(3, len(ic_sku)):##copyni
        if pd.isna(ic_sku.iat[idx, a_col_sku]) or ic_sku.iat[idx, a_col_sku] == "":
            c = 1
            d += 1
        else:
            ic_sku.at[idx, col_c] = c
            ic_sku.at[idx, col_d] = d
            c += 1

    # 3) Delete rows where A is blank
    ic_sku = ic_sku[~ic_sku.iloc[:, a_col_sku].isna()].reset_index(drop=True)

    # ==================================================
    # === Repeat count (R,S) on IC MAIN like VBA R/S ===
    # ==================================================
    col_r = "RepeatCount"
    col_s = "CumulativeOffset"
    ic_main[col_r] = 0
    ic_main[col_s] = 0

    # Count how many SKU rows per group (SKU_group)
    sku_group_counts = (
        ic_sku[col_d].value_counts().to_dict()
        if col_d in ic_sku.columns
        else {}
    )

    for idx in range(0, len(ic_main)):#copyni
        q_val = ic_main.at[idx, col_q]
        count = int(sku_group_counts.get(q_val, 0))
        ic_main.at[idx, col_r] = count

        if idx == 0:
            ic_main.at[idx, col_s] = 2
        else:
            prev_r = ic_main.at[idx - 1, col_r]
            prev_s = ic_main.at[idx - 1, col_s]
            ic_main.at[idx, col_s] = prev_r + prev_s

    # ==================================================
    # === XLOOKUP-style mapping for IC SKU (E column) ===
    # ==================================================
    col_e = "SKU_row_no"
    ic_sku[col_e] = pd.NA

    lookup = (
        ic_main[[col_q, col_s]]
        .dropna()
        .drop_duplicates(subset=[col_q])
        .set_index(col_q)[col_s]
        .to_dict()
    )

    prev_d_val = None
    prev_e_val = None
    for idx in range(0, len(ic_sku)):#copyni
        d_val = ic_sku.at[idx, col_d]
        if pd.isna(d_val):
            continue

        base = lookup.get(d_val, pd.NA)
        if idx > 1 and d_val == prev_d_val and prev_e_val is not None:
            ic_sku.at[idx, col_e] = prev_e_val + 1
        else:
            ic_sku.at[idx, col_e] = base

        prev_d_val = d_val
        prev_e_val = ic_sku.at[idx, col_e]

    # ==================================================
    # ===== Final U:Z-style summary table (IC MAIN) =====
    # ==================================================
    # Equivalent of:
    #  No | Scheme ID(F) | Scheme Promo No(B) | Desc(D) | Period From(L) | Period To(M)
    out_cols = [
        "No",
        "SchemeID",
        "SchemePromotionNumber",
        "SchemeDescription",
        "PeriodFrom",
        "PeriodTo",
    ]
    ic_main_out = pd.DataFrame(
        index=range(max(0, len(ic_main) - 1)), columns=out_cols
    )

    numb = 1
    for i_out, src_idx in enumerate(range(1, len(ic_main))):
        def safe(col_idx):
            return ic_main.iat[src_idx, col_idx] if col_idx < ic_main.shape[1] else pd.NA

        ic_main_out.iat[i_out, 0] = numb         # No
        ic_main_out.iat[i_out, 1] = safe(5)      # F
        ic_main_out.iat[i_out, 2] = safe(1)      # B
        ic_main_out.iat[i_out, 3] = safe(3)      # D
        ic_main_out.iat[i_out, 4] = safe(11)     # L
        ic_main_out.iat[i_out, 5] = safe(12)     # M
        numb += 1

    return ic_main, ic_sku, ic_main_out

# simple in-memory cache for the current session run
PROMO_CACHE = {}

@app.route('/api/promotions/auto/import', methods=['POST'])
def import_monthly_ic_promotion():
    uploaded = request.files.get('file')
    if not uploaded:
        return jsonify({"error": "No file uploaded"}), 400

    ic_main, ic_sku, ic_main_out = process_monthly_ic_promotion(uploaded)

    # store temporarily in memory (Pandas "database")
    PROMO_CACHE['ic_main'] = ic_main
    PROMO_CACHE['ic_sku'] = ic_sku
    PROMO_CACHE['ic_main_out'] = ic_main_out

    print("summary IC:", ic_sku)
    # send only a preview back to frontend
    return jsonify({
        "message": "File imported and processed successfully.",
        "summary_rows": len(ic_main_out),
        "summary_preview": ic_main_out.head(50).to_dict(orient="records"),
    })
##----------------EXCEL PROMO AUTO EXPORT-------------------
IC_TEMPLATE_COLUMNS = [
    "PromotionCode",
    "PromotionDescription",
    "PromotionType",
    "NationalBudget",
    "TestScheme",
    "BuyBase",
    "GetBase",
    "MultiplicationFactor",
    "StartDate",
    "EndDate",
    "PromotionStatus",
    "PromotionQuotaLevel",
    "PromotionQuotaOn",
    "PromotionClaimable",
    "OPSOID",
    "MaxInvoicesperOutlet",
    "MinBuySKUs",
    "PromotionUOM",
    "AlternatePromotionDescription",
    "UserExpire",
    "PromotionSlab",
    "PromotionSlabDescription",
    "RangeLow",
    "RangeHigh",
    "PromotionReturn",
    "ForEvery",
    "PurchaseLimit",
    "ProductHierarchyLevel",
    "ProductHierarchyCode",
    "Exclude",
    "ConditionGroup",
    "GroupType",
    "MinimumQty",
    "BasketPromotion",
    "CriteriaType",
    "CriteriaValue",
    "CriteriaExclude",
]

def build_ic_template(ic_main: pd.DataFrame, ic_sku: pd.DataFrame) -> pd.DataFrame:
    """
    Python translation of your CommandButton31_Click logic.
    Uses ic_main + ic_sku (already processed in-memory) to build IC Template rows.
    """

    if ic_main is None or ic_sku is None:
        raise ValueError("ic_main or ic_sku is missing. Run import/process first.")

    rows = []

    # SAFETY: helper to read by Excel-like column index if exists
    def col(df, idx):
        return df.iloc[:, idx] if idx < df.shape[1] else pd.Series([pd.NA] * len(df))

    # Loops: For i = 2 To lastRow: For j = 1 To R(i)
    # Data rows in Pandas: index 1..end (since VBA starts at row2)
    for i in range(0, len(ic_main)):
        # RepeatCount (VBA R-column)
        repeat = ic_main.get("RepeatCount", pd.Series([1] * len(ic_main))).iloc[i]
        try:
            repeat = int(repeat)
        except (TypeError, ValueError):
            repeat = 1
        if repeat < 1:
            continue

        promo_code = col(ic_main, 1).iat[i]     # B
        desc = col(ic_main, 3).iat[i]           # D
        ops_oid = col(ic_main, 5).iat[i]        # F
        uom = col(ic_main, 8).iat[i]            # I
        range_low = col(ic_main, 9).iat[i]      # J
        promo_return = col(ic_main, 10).iat[i]  # K
        start_date = col(ic_main, 11).iat[i]    # L
        end_date = col(ic_main, 12).iat[i]      # M
        crit_type = col(ic_main, 13).iat[i]     # N
        crit_value = col(ic_main, 14).iat[i]    # O

        for _ in range(repeat):
            row = {
                "PromotionCode": promo_code,
                "PromotionDescription": desc,
                "PromotionType": "T",
                "NationalBudget": 999999,
                "TestScheme": 1,
                "BuyBase": 3,
                "GetBase": 5,
                "MultiplicationFactor": 1,
                "StartDate": start_date,
                "EndDate": end_date,
                "PromotionStatus": 1,
                "PromotionQuotaLevel": "SR",
                "PromotionQuotaOn": "D",
                "PromotionClaimable": 1,
                "OPSOID": ops_oid,
                "MaxInvoicesperOutlet": 99999,
                "MinBuySKUs": 0,
                "PromotionUOM": uom,
                "AlternatePromotionDescription": "AlternatePromotionDescription",
                "UserExpire": 0,
                "PromotionSlab": 1,
                "PromotionSlabDescription": "PromotionSlab",
                "RangeLow": range_low,
                "RangeHigh": 999999,
                "PromotionReturn": promo_return,
                "ForEvery": range_low,
                "PurchaseLimit": 0,
                "ProductHierarchyLevel": "S",
                "ProductHierarchyCode": None,  # fill later using IC SKU
                "Exclude": 0,
                "ConditionGroup": 1,
                "GroupType": "Q",
                "MinimumQty": 0,
                "BasketPromotion": 1,
                "CriteriaType": crit_type,
                "CriteriaValue": crit_value,
                "CriteriaExclude": 0,
            }
            rows.append(row)

    ic_template = pd.DataFrame(rows, columns=IC_TEMPLATE_COLUMNS)

    # ========= Fill ProductHierarchyCode (AC column) from IC SKU =========
    # VBA logic:
    #   For each SKU row k:
    #     group = SKU_group
    #     totalSKU = count of rows in IC SKU with that group
    #     repeatCountForGroup = CountIf(IC Main Q:Q, group)
    #     baseRow = SKU_row_no (E)
    #     AC[ baseRow + totalSKU*(l-1) ] = SKU code
    #
    # We reproduce assuming:
    #   - ic_sku['SKU_group']
    #   - ic_sku['SKU_row_no']
    #   - first column of ic_sku is SKU code

    if (
        "SKU_group" in ic_sku.columns
        and "SKU_row_no" in ic_sku.columns
        and len(ic_template) > 0
    ):
        # For each group, know how many slabs/promos exist
        group_counts_main = (
            ic_main.get("Q_group")
            .value_counts()
            .to_dict()
            if "Q_group" in ic_main.columns
            else {}
        )

        for _, sku_row in ic_sku.iterrows():
            group_id = sku_row.get("SKU_group")
            if pd.isna(group_id):
                continue

            sku_code = sku_row.iloc[0]
            total_sku = int((ic_sku["SKU_group"] == group_id).sum())
            repeat_for_group = int(group_counts_main.get(group_id, 0))
            if total_sku == 0 or repeat_for_group == 0:
                continue

            base_row_excel = int(sku_row.get("SKU_row_no", 0))
            if base_row_excel <= 1:
                continue

            # Translate Excel row to DataFrame index: row2 -> idx0
            for l in range(0, repeat_for_group + 1):
                target_excel_row = base_row_excel + (total_sku * (l - 1))
                target_idx = target_excel_row - 2
                if 0 <= target_idx < len(ic_template):
                    ic_template.at[target_idx, "ProductHierarchyCode"] = sku_code

    # Format dates as dd/mm/yyyy strings (like your requirement)
    for col_name in ["StartDate", "EndDate"]:
        if col_name in ic_template.columns:
            ic_template[col_name] = ic_template[col_name].apply(_format_ddmmyyyy)

    # PromotionReturn (Y) should be numeric with 2 decimals (we keep as float; Excel formatting done later)
    if "PromotionReturn" in ic_template.columns:
        ic_template["PromotionReturn"] = pd.to_numeric(
            ic_template["PromotionReturn"], errors="coerce"
        )

    return ic_template


def _format_ddmmyyyy(val):
    if pd.isna(val) or val == "":
        return ""
    try:
        d = pd.to_datetime(val)
    except Exception:
        return str(val)
    return d.strftime("%d/%m/%Y")


def build_ic_template_excel_bytes(ic_main: pd.DataFrame, ic_sku: pd.DataFrame) -> BytesIO:
    """
    Build IC Template and return as an in-memory .xlsx file (BytesIO).
    """
    df = build_ic_template(ic_main, ic_sku)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="IC Template")

        # Optional styling similar to VBA:
        workbook = writer.book
        ws = writer.sheets["IC Template"]

        # Date columns I,J (StartDate, EndDate) are already formatted as text dd/mm/yyyy,
        # but if you prefer real dates, adjust logic above.

        # PromotionReturn (Y) with 2 decimals if numeric
        num_fmt = workbook.add_format({"num_format": "0.00"})
        # Column index: 0-based; PromotionReturn is at position 24
        ws.set_column(24, 24, 12, num_fmt)  # col Y

    output.seek(0)
    return output

@app.route('/api/promotions/auto/export', methods=['GET'])
def export_ic_template():
    ic_main = PROMO_CACHE.get("ic_main")
    ic_sku = PROMO_CACHE.get("ic_sku")

    if ic_main is None or ic_sku is None:
        return jsonify({"error": "No processed data found. Please import & generate first."}), 400

    output = build_ic_template_excel_bytes(ic_main, ic_sku)
    return send_file(
        output,
        as_attachment=True,
        download_name="IC_Promo_Template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.get("/test")
def api_test():
    return jsonify(status="ok"), 200

@app.get("/")
def api_root():
    return jsonify(status="ok", service="mdm-backend"), 200

@app.get("/__deps")
def __deps():
    import sys
    import pandas as pd
    out = {
        "python": sys.version,
        "executable": sys.executable,
        "pandas": pd.__version__,
    }
    try:
        import xlrd
        out["xlrd"] = xlrd.__version__
    except Exception as e:
        out["xlrd"] = f"missing: {e}"
    try:
        import openpyxl
        out["openpyxl"] = openpyxl.__version__
    except Exception as e:
        out["openpyxl"] = f"missing: {e}"
    return out

##---------------email tracker---------------------


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--port", type=int, default=int(os.getenv("PORT", 5000)))
    parser.add_argument("--creds", type=str, default=None)
    args = parser.parse_args()

    # ✅ Initialize once here
    init_sheets(args.creds)

    app.run(host="127.0.0.1", port=args.port, debug=False, use_reloader=False)

